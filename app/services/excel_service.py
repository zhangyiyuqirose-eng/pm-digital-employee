"""
PM Digital Employee - Excel Service
项目经理数字员工系统 - Excel导入导出服务

提供Excel模板生成、数据导出、数据解析、数据导入功能。
支持项目、任务、里程碑、风险、成本五个模块。
"""

import os
import uuid
import json
from datetime import datetime, date, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging import get_logger
from app.core.validation_config import get_module_config, get_all_module_names
from app.domain.enums import ImportMode
from app.domain.models.excel_import_log import ExcelImportLog
from app.services.validation_service import ValidationService

logger = get_logger(__name__)

# Excel模板版本号
TEMPLATE_VERSION = "v1.2.0"

# 样式定义
HEADER_FONT = Font(bold=True, size=11)
REQUIRED_HEADER_FONT = Font(bold=True, size=11, color="FF0000")  # 红色表示必填
EXAMPLE_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # 灰色背景
TITLE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 蓝色标题
TITLE_FONT = Font(bold=True, size=12, color="FFFFFF")
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExcelService:
    """
    Excel导入导出服务.

    提供模板生成、数据导出、数据解析、数据导入功能。
    """

    def __init__(self, session: AsyncSession) -> None:
        """
        初始化Excel服务.

        Args:
            session: 数据库会话
        """
        self.session = session
        self.validation_service = ValidationService()

    # ==================== 模板生成 ====================

    def generate_template(self, module: str) -> BytesIO:
        """
        生成Excel模板.

        每个模板包含两个工作表：
        1. 说明页：字段说明、填写指南、示例
        2. 数据页：模板版本、列标题、示例数据

        Args:
            module: 模块名称（project/task/milestone/risk/cost）

        Returns:
            BytesIO: Excel文件流

        Raises:
            ValueError: 模块不存在
        """
        config = get_module_config(module)
        if not config:
            raise ValueError(f"模块 '{module}' 不存在，支持的模块: {get_all_module_names()}")

        # 创建Workbook
        wb = Workbook()

        # 创建说明页
        ws_instruction = wb.active
        ws_instruction.title = "说明"
        self._build_instruction_sheet(ws_instruction, config)

        # 创建数据页
        ws_data = wb.create_sheet(title="数据")
        self._build_data_sheet(ws_data, config)

        # 写入BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f"Generated Excel template for module: {module}")
        return buffer

    def _build_instruction_sheet(
        self,
        ws: Any,
        config: Any,
    ) -> None:
        """
        构建说明页.

        包含字段说明、填写指南、示例。

        Args:
            ws: 工作表对象
            config: 模块校验配置
        """
        # 标题行
        ws.merge_cells("A1:F1")
        ws["A1"] = f"{config.display_name}数据导入模板 - 填写说明"
        ws["A1"].font = TITLE_FONT
        ws["A1"].fill = TITLE_FILL
        ws["A1"].alignment = Alignment(horizontal="center")

        # 版本信息
        ws["A2"] = "模板版本："
        ws["B2"] = TEMPLATE_VERSION
        ws["C2"] = "更新时间："
        ws["D2"] = datetime.now().strftime("%Y-%m-%d")

        # 空行
        ws["A3"] = ""

        # 字段说明标题
        ws["A4"] = "字段说明"
        ws["A4"].font = HEADER_FONT

        # 表头
        headers = ["字段名称", "显示名称", "数据类型", "是否必填", "允许值/范围", "填写说明"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = HEADER_FONT
            cell.border = BORDER_THIN

        # 字段详情
        for row_idx, field in enumerate(config.fields, 6):
            ws.cell(row=row_idx, column=1, value=field.field_name).border = BORDER_THIN
            ws.cell(row=row_idx, column=2, value=field.display_name).border = BORDER_THIN
            ws.cell(row=row_idx, column=3, value=field.field_type).border = BORDER_THIN

            required_text = "是" if field.required else "否"
            ws.cell(row=row_idx, column=4, value=required_text).border = BORDER_THIN

            # 允许值/范围
            allowed_values = ""
            if field.enum_values:
                allowed_values = "/".join(field.enum_values)
            elif field.min_value is not None or field.max_value is not None:
                min_v = field.min_value if field.min_value is not None else ""
                max_v = field.max_value if field.max_value is not None else ""
                allowed_values = f"{min_v} ~ {max_v}"
            elif field.min_length is not None or field.max_length is not None:
                min_l = field.min_length if field.min_length is not None else ""
                max_l = field.max_length if field.max_length is not None else ""
                allowed_values = f"长度 {min_l} ~ {max_l}"
            ws.cell(row=row_idx, column=5, value=allowed_values).border = BORDER_THIN

            # 填写说明
            description = ""
            if field.field_type == "date":
                description = "格式：YYYY-MM-DD"
            elif field.field_type == "datetime":
                description = "格式：YYYY-MM-DD HH:MM:SS"
            elif field.field_type == "enum":
                description = "请从下拉列表选择"
            elif field.required:
                description = "必填字段，不能为空"
            ws.cell(row=row_idx, column=6, value=description).border = BORDER_THIN

        # 设置列宽
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 25

    def _build_data_sheet(
        self,
        ws: Any,
        config: Any,
    ) -> None:
        """
        构建数据页.

        包含模板版本、列标题、示例数据、下拉选择框。

        Args:
            ws: 工作表对象
            config: 模块校验配置
        """
        # 第一行：模板版本、更新时间
        ws["A1"] = f"模板版本: {TEMPLATE_VERSION}"
        ws["B1"] = f"更新时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A1"].font = Font(italic=True, size=9)
        ws["B1"].font = Font(italic=True, size=9)

        # 第二行：列标题
        for col, field in enumerate(config.fields, 1):
            cell = ws.cell(row=2, column=col, value=field.display_name)

            # 必填字段红色标记
            if field.required:
                cell.font = REQUIRED_HEADER_FONT
            else:
                cell.font = HEADER_FONT

            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center")

            # 设置列宽
            col_letter = get_column_letter(col)
            if field.field_type == "str":
                ws.column_dimensions[col_letter].width = max(15, field.max_length // 2 if field.max_length else 20)
            else:
                ws.column_dimensions[col_letter].width = 12

        # 第三行：示例数据（灰色背景）
        for col, field in enumerate(config.fields, 1):
            example_value = self._get_example_value(field)
            cell = ws.cell(row=3, column=col, value=example_value)
            cell.fill = EXAMPLE_FILL
            cell.border = BORDER_THIN

        # 添加枚举下拉选择框
        for col, field in enumerate(config.fields, 1):
            if field.enum_values:
                col_letter = get_column_letter(col)
                dv = DataValidation(
                    type="list",
                    formula1=f'"{"，".join(field.enum_values)}"',
                    allow_blank=not field.required,
                )
                dv.error = "请从下拉列表中选择正确的值"
                dv.errorTitle = "无效输入"
                dv.prompt = f"请选择 {field.display_name}"
                dv.promptTitle = "选择提示"
                ws.add_data_validation(dv)
                # 从第4行开始应用（跳过版本行、标题行、示例行）
                dv.add(f"{col_letter}4:{col_letter}1000")

    def _get_example_value(self, field: Any) -> str:
        """
        获取字段示例值.

        Args:
            field: 字段校验配置

        Returns:
            str: 示例值
        """
        if field.enum_values:
            return field.enum_values[0]
        elif field.field_type == "str":
            if field.field_name == "name":
                return f"示例{field.display_name}"
            elif field.field_name == "project_id":
                return "项目ID示例"
            return "示例文本"
        elif field.field_type == "int":
            if "progress" in field.field_name:
                return "50"
            return "100"
        elif field.field_type == "float":
            return "1000.00"
        elif field.field_type == "date":
            return datetime.now().strftime("%Y-%m-%d")
        elif field.field_type == "datetime":
            return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        elif field.field_type == "bool":
            return "是"
        return ""

    # ==================== 数据导出 ====================

    async def export_data(
        self,
        module: str,
        project_id: Optional[uuid.UUID] = None,
    ) -> BytesIO:
        """
        导出数据为Excel.

        根据模块和项目ID导出对应数据。

        Args:
            module: 模块名称
            project_id: 项目ID（可选，部分模块需要）

        Returns:
            BytesIO: Excel文件流

        Raises:
            ValueError: 模块不存在或缺少必要参数
        """
        config = get_module_config(module)
        if not config:
            raise ValueError(f"模块 '{module}' 不存在")

        # 获取数据
        data_list = await self._fetch_module_data(module, project_id)

        # 创建Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "数据"

        # 写入导出时间
        ws["A1"] = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["B1"] = f"数据条数: {len(data_list)}"

        # 写入列标题
        for col, field in enumerate(config.fields, 1):
            cell = ws.cell(row=2, column=col, value=field.display_name)
            cell.font = HEADER_FONT
            cell.border = BORDER_THIN

        # 写入数据
        for row_idx, data in enumerate(data_list, 3):
            for col, field in enumerate(config.fields, 1):
                value = self._format_export_value(data.get(field.field_name), field)
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = BORDER_THIN

        # 设置列宽
        for col, field in enumerate(config.fields, 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15

        # 写入BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f"Exported Excel data for module: {module}, count: {len(data_list)}")
        return buffer

    async def _fetch_module_data(
        self,
        module: str,
        project_id: Optional[uuid.UUID] = None,
    ) -> List[Dict[str, Any]]:
        """
        从数据库获取模块数据.

        Args:
            module: 模块名称
            project_id: 项目ID

        Returns:
            List[Dict]: 数据列表
        """
        # 根据模块调用对应的Repository/Service
        if module == "project":
            from app.repositories.project_repository import ProjectRepository
            repo = ProjectRepository(self.session)
            if project_id:
                project = await repo.get_by_id(project_id)
                return [self._model_to_dict(project)] if project else []
            else:
                projects = await repo.get_all(limit=1000)
                return [self._model_to_dict(p) for p in projects]

        elif module == "task":
            from app.repositories.task_repository import TaskRepository
            repo = TaskRepository(self.session)
            if project_id:
                tasks = await repo.list_by_project(project_id, limit=1000)
                return [self._model_to_dict(t) for t in tasks]
            else:
                tasks = await repo.get_all(limit=1000)
                return [self._model_to_dict(t) for t in tasks]

        elif module == "milestone":
            from app.repositories.milestone_repository import MilestoneRepository
            repo = MilestoneRepository(self.session)
            if project_id:
                milestones = await repo.list_by_project(project_id, limit=1000)
                return [self._model_to_dict(m) for m in milestones]
            else:
                milestones = await repo.get_all(limit=1000)
                return [self._model_to_dict(m) for m in milestones]

        elif module == "risk":
            from app.repositories.risk_repository import RiskRepository
            repo = RiskRepository(self.session)
            if project_id:
                risks = await repo.list_by_project(project_id, limit=1000)
                return [self._model_to_dict(r) for r in risks]
            else:
                risks = await repo.get_all(limit=1000)
                return [self._model_to_dict(r) for r in risks]

        elif module == "cost":
            from app.repositories.cost_repository import CostRepository
            repo = CostRepository(self.session)
            if project_id:
                costs = await repo.list_by_project(project_id, limit=1000)
                return [self._model_to_dict(c) for c in costs]
            else:
                costs = await repo.get_all(limit=1000)
                return [self._model_to_dict(c) for c in costs]

        return []

    def _model_to_dict(self, model: Any) -> Dict[str, Any]:
        """
        将ORM模型转换为字典.

        Args:
            model: ORM模型实例

        Returns:
            Dict: 字典数据
        """
        result = {}
        for column in model.__table__.columns:
            value = getattr(model, column.name)
            result[column.name] = value
        return result

    def _format_export_value(self, value: Any, field: Any) -> str:
        """
        格式化导出值.

        Args:
            value: 原始值
            field: 字段配置

        Returns:
            str: 格式化后的值
        """
        if value is None:
            return ""

        if isinstance(value, uuid.UUID):
            return str(value)

        if isinstance(value, (date, datetime)):
            return value.strftime("%Y-%m-%d")

        if isinstance(value, Decimal):
            return float(value)

        if isinstance(value, bool):
            return "是" if value else "否"

        return str(value)

    # ==================== 数据解析 ====================

    def parse_excel(
        self,
        file_path: str,
        module: str,
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        解析Excel数据.

        读取Excel文件并转换为数据字典列表。

        Args:
            file_path: Excel文件路径
            module: 模块名称

        Returns:
            Tuple[List[Dict], List[Dict]]: (有效数据列表, 行级错误列表)
        """
        config = get_module_config(module)
        if not config:
            raise ValueError(f"模块 '{module}' 不存在")

        # 加载Excel文件
        wb = load_workbook(file_path, data_only=True)
        ws = wb["数据"] if "数据" in wb.sheetnames else wb.active

        # 获取字段映射（列标题 -> 字段名）
        field_mapping = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=2, column=col).value
            if header:
                for field in config.fields:
                    if field.display_name == header:
                        field_mapping[col] = field.field_name
                        break

        # 解析数据行（从第4行开始，跳过版本、标题、示例）
        data_list: List[Dict[str, Any]] = []
        row_errors: List[Dict[str, Any]] = []

        for row_idx in range(4, ws.max_row + 1):
            row_data = {}
            is_row_valid = True

            for col in range(1, ws.max_column + 1):
                field_name = field_mapping.get(col)
                if not field_name:
                    continue

                value = ws.cell(row=row_idx, column=col).value

                # 空行跳过
                if value is None and col == 1:
                    is_row_valid = False
                    break

                row_data[field_name] = value

            if is_row_valid and row_data:
                data_list.append(row_data)

        logger.info(f"Parsed Excel file: {file_path}, module: {module}, rows: {len(data_list)}")
        return data_list, row_errors

    # ==================== 数据导入 ====================

    async def import_data(
        self,
        data_list: List[Dict[str, Any]],
        module: str,
        import_mode: str,
        project_id: Optional[uuid.UUID] = None,
        user_id: Optional[str] = None,
    ) -> ExcelImportLog:
        """
        批量导入数据.

        根据导入模式执行数据导入：
        - full_replace: 全量替换（删除旧数据，插入新数据）
        - incremental_update: 增量更新（根据ID判断，存在则更新，不存在则插入）
        - append_only: 仅追加（只插入新数据）

        Args:
            data_list: 数据列表
            module: 模块名称
            import_mode: 导入模式
            project_id: 项目ID（可选）
            user_id: 用户ID（可选）

        Returns:
            ExcelImportLog: 导入日志
        """
        # 创建导入日志
        import_log = ExcelImportLog(
            file_name=f"{module}_import_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
            import_mode=import_mode,
            template_version=TEMPLATE_VERSION,
            rows_total=len(data_list),
        )
        self.session.add(import_log)
        await self.session.flush()

        # 1. 执行数据校验
        validation_results = self.validation_service.validate_batch(data_list, module)

        # 统计校验结果
        valid_data = []
        row_errors = []

        for idx, result in enumerate(validation_results):
            if result.is_valid:
                valid_data.append(result.validated_data)
            else:
                row_errors.append({
                    "row_index": idx + 1,
                    "errors": result.errors,
                })

        # 更新导入日志校验结果
        import_log.validation_passed = len(row_errors) == 0
        import_log.validation_errors = json.dumps(row_errors[:100]) if row_errors else None  # 限制存储数量
        import_log.rows_failed = len(row_errors)

        # 2. 执行数据导入（仅处理校验通过的数据）
        if valid_data:
            import_result = await self._execute_import(
                valid_data,
                module,
                import_mode,
                project_id,
            )
            import_log.rows_imported = import_result.get("inserted", 0)
            import_log.rows_updated = import_result.get("updated", 0)
            import_log.rows_skipped = import_result.get("skipped", 0)

        # 更新行级错误详情
        import_log.row_errors = json.dumps(row_errors[:500]) if row_errors else None  # 限制存储数量

        await self.session.flush()
        await self.session.refresh(import_log)

        logger.info(
            f"Import completed: module={module}, mode={import_mode}, "
            f"total={import_log.rows_total}, imported={import_log.rows_imported}, "
            f"updated={import_log.rows_updated}, failed={import_log.rows_failed}"
        )

        return import_log

    async def _execute_import(
        self,
        data_list: List[Dict[str, Any]],
        module: str,
        import_mode: str,
        project_id: Optional[uuid.UUID] = None,
    ) -> Dict[str, int]:
        """
        执行数据导入操作.

        Args:
            data_list: 校验通过的数据列表
            module: 模块名称
            import_mode: 导入模式
            project_id: 项目ID

        Returns:
            Dict: 导入统计 {inserted, updated, skipped}
        """
        result = {"inserted": 0, "updated": 0, "skipped": 0}

        if module == "project":
            from app.repositories.project_repository import ProjectRepository
            repo = ProjectRepository(self.session)

            if import_mode == ImportMode.FULL_REPLACE.value:
                # 全量替换：清空旧数据
                await self._clear_project_data(project_id)
                for data in data_list:
                    await repo.create(data)
                    result["inserted"] += 1

            elif import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                for data in data_list:
                    existing_id = data.get("id")
                    if existing_id:
                        try:
                            existing_id = uuid.UUID(existing_id)
                            await repo.update(existing_id, data)
                            result["updated"] += 1
                        except Exception:
                            await repo.create(data)
                            result["inserted"] += 1
                    else:
                        await repo.create(data)
                        result["inserted"] += 1

            elif import_mode == ImportMode.APPEND_ONLY.value:
                for data in data_list:
                    await repo.create(data)
                    result["inserted"] += 1

        elif module == "task":
            from app.repositories.task_repository import TaskRepository
            repo = TaskRepository(self.session)

            # 任务必须有project_id
            target_project_id = project_id or data_list[0].get("project_id")
            if not target_project_id:
                raise ValueError("任务导入需要指定项目ID")

            if import_mode == ImportMode.FULL_REPLACE.value:
                await self._clear_tasks(target_project_id)
                for data in data_list:
                    data["project_id"] = target_project_id
                    await repo.create(data)
                    result["inserted"] += 1

            elif import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                for data in data_list:
                    data["project_id"] = target_project_id
                    existing_id = data.get("id")
                    if existing_id:
                        try:
                            existing_id = uuid.UUID(existing_id)
                            await repo.update_in_project(existing_id, target_project_id, data)
                            result["updated"] += 1
                        except Exception:
                            await repo.create(data)
                            result["inserted"] += 1
                    else:
                        await repo.create(data)
                        result["inserted"] += 1

            elif import_mode == ImportMode.APPEND_ONLY.value:
                for data in data_list:
                    data["project_id"] = target_project_id
                    await repo.create(data)
                    result["inserted"] += 1

        elif module == "milestone":
            from app.repositories.milestone_repository import MilestoneRepository
            repo = MilestoneRepository(self.session)

            target_project_id = project_id or data_list[0].get("project_id")
            if not target_project_id:
                raise ValueError("里程碑导入需要指定项目ID")

            if import_mode == ImportMode.FULL_REPLACE.value:
                await self._clear_milestones(target_project_id)
                for data in data_list:
                    data["project_id"] = target_project_id
                    await repo.create(data)
                    result["inserted"] += 1

            elif import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                for data in data_list:
                    data["project_id"] = target_project_id
                    existing_id = data.get("id")
                    if existing_id:
                        try:
                            existing_id = uuid.UUID(existing_id)
                            await repo.update_in_project(existing_id, target_project_id, data)
                            result["updated"] += 1
                        except Exception:
                            await repo.create(data)
                            result["inserted"] += 1
                    else:
                        await repo.create(data)
                        result["inserted"] += 1

            elif import_mode == ImportMode.APPEND_ONLY.value:
                for data in data_list:
                    data["project_id"] = target_project_id
                    await repo.create(data)
                    result["inserted"] += 1

        elif module == "risk":
            from app.repositories.risk_repository import RiskRepository
            repo = RiskRepository(self.session)

            target_project_id = project_id or data_list[0].get("project_id")
            if not target_project_id:
                raise ValueError("风险导入需要指定项目ID")

            if import_mode == ImportMode.FULL_REPLACE.value:
                await self._clear_risks(target_project_id)
                for data in data_list:
                    data["project_id"] = target_project_id
                    await repo.create(data)
                    result["inserted"] += 1

            elif import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                for data in data_list:
                    data["project_id"] = target_project_id
                    existing_id = data.get("id")
                    if existing_id:
                        try:
                            existing_id = uuid.UUID(existing_id)
                            await repo.update_in_project(existing_id, target_project_id, data)
                            result["updated"] += 1
                        except Exception:
                            await repo.create(data)
                            result["inserted"] += 1
                    else:
                        await repo.create(data)
                        result["inserted"] += 1

            elif import_mode == ImportMode.APPEND_ONLY.value:
                for data in data_list:
                    data["project_id"] = target_project_id
                    await repo.create(data)
                    result["inserted"] += 1

        elif module == "cost":
            from app.repositories.cost_repository import CostRepository
            repo = CostRepository(self.session)

            target_project_id = project_id or data_list[0].get("project_id")
            if not target_project_id:
                raise ValueError("成本导入需要指定项目ID")

            if import_mode == ImportMode.FULL_REPLACE.value:
                await self._clear_costs(target_project_id)
                for data in data_list:
                    data["project_id"] = target_project_id
                    await repo.create(data)
                    result["inserted"] += 1

            elif import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                for data in data_list:
                    data["project_id"] = target_project_id
                    existing_id = data.get("id")
                    if existing_id:
                        try:
                            existing_id = uuid.UUID(existing_id)
                            await repo.update_in_project(existing_id, target_project_id, data)
                            result["updated"] += 1
                        except Exception:
                            await repo.create(data)
                            result["inserted"] += 1
                    else:
                        await repo.create(data)
                        result["inserted"] += 1

            elif import_mode == ImportMode.APPEND_ONLY.value:
                for data in data_list:
                    data["project_id"] = target_project_id
                    await repo.create(data)
                    result["inserted"] += 1

        await self.session.flush()
        return result

    async def _clear_project_data(self, project_id: uuid.UUID) -> None:
        """
        清空项目相关数据.

        Args:
            project_id: 项目ID
        """
        # 删除任务、里程碑、风险、成本
        await self._clear_tasks(project_id)
        await self._clear_milestones(project_id)
        await self._clear_risks(project_id)
        await self._clear_costs(project_id)

    async def _clear_tasks(self, project_id: uuid.UUID) -> None:
        """删除项目下的所有任务."""
        from sqlalchemy import delete
        from app.domain.models.task import Task
        await self.session.execute(delete(Task).where(Task.project_id == project_id))

    async def _clear_milestones(self, project_id: uuid.UUID) -> None:
        """删除项目下的所有里程碑."""
        from sqlalchemy import delete
        from app.domain.models.milestone import Milestone
        await self.session.execute(delete(Milestone).where(Milestone.project_id == project_id))

    async def _clear_risks(self, project_id: uuid.UUID) -> None:
        """删除项目下的所有风险."""
        from sqlalchemy import delete
        from app.domain.models.risk import ProjectRisk
        await self.session.execute(delete(ProjectRisk).where(ProjectRisk.project_id == project_id))

    async def _clear_costs(self, project_id: uuid.UUID) -> None:
        """删除项目下的所有成本."""
        from sqlalchemy import delete
        from app.domain.models.cost import ProjectCostBudget, ProjectCostActual
        await self.session.execute(delete(ProjectCostBudget).where(ProjectCostBudget.project_id == project_id))
        await self.session.execute(delete(ProjectCostActual).where(ProjectCostActual.project_id == project_id))

    # ==================== 导入日志查询 ====================

    async def get_import_logs(
        self,
        module: Optional[str] = None,
        limit: int = 50,
    ) -> List[ExcelImportLog]:
        """
        获取导入日志列表.

        Args:
            module: 模块名称过滤（可选）
            limit: 返回数量限制

        Returns:
            List[ExcelImportLog]: 导入日志列表
        """
        from sqlalchemy import select

        query = select(ExcelImportLog).order_by(ExcelImportLog.created_at.desc()).limit(limit)
        result = await self.session.execute(query)
        return list(result.scalars().all())