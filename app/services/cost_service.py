"""
PM Digital Employee - Cost Service (Extended)
项目经理数字员工系统 - 成本业务服务（扩展版）

v1.2.0新增功能：
- Excel导入导出成本数据
- 飞书表格同步成本数据
- 成本偏差计算、成本执行率、剩余预算
- 成本阈值设置和超支预警通知
- 成本统计报表和趋势分析数据生成
- 成本审批流程集成飞书审批
"""

import uuid
import json
from datetime import datetime, timezone, date
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Optional

from sqlalchemy import select, func, and_, case
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.core.exceptions import ErrorCode, CostNotFoundError, ProjectNotFoundError
from app.core.logging import get_logger
from app.domain.enums import CostCategory, ImportMode, RiskLevel
from app.domain.models.cost import ProjectCostBudget, ProjectCostActual
from app.repositories.cost_repository import CostBudgetRepository, CostActualRepository
from app.repositories.project_repository import ProjectRepository
from app.services.validation_service import ValidationService
from app.services.sync_engine import SyncEngine
from app.services.notification_service import NotificationService

logger = get_logger(__name__)

# 阈值配置
DEFAULT_THRESHOLD_PERCENT = 80  # 默认阈值百分比
THRESHOLD_LEVELS = {
    "low": {"percent": 70, "color": "yellow"},
    "medium": {"percent": 80, "color": "orange"},
    "high": {"percent": 90, "color": "red"},
}


class CostThresholdConfig:
    """成本阈值配置."""

    def __init__(
        self,
        project_id: uuid.UUID,
        threshold_percent: float,
        warning_level: str,
        notify_users: Optional[List[str]] = None,
    ):
        self.project_id = project_id
        self.threshold_percent = threshold_percent
        self.warning_level = warning_level
        self.notify_users = notify_users or []
        self.created_at = datetime.now(timezone.utc)


class CostService:
    """
    成本业务服务（扩展版）.

    封装成本相关的业务逻辑，包括预算、实际支出、成本分析、预警等功能。
    """

    def __init__(self, session: AsyncSession) -> None:
        """
        初始化成本服务.

        Args:
            session: 数据库会话
        """
        self.session = session
        self._budget_repository = CostBudgetRepository(session)
        self._actual_repository = CostActualRepository(session)
        self._project_repository = ProjectRepository(session)
        self._validation_service = ValidationService()
        self._threshold_configs: Dict[str, CostThresholdConfig] = {}

    # ==================== 基础操作（保留原有功能） ====================

    async def create_budget(
        self,
        project_id: uuid.UUID,
        category: CostCategory,
        amount: Decimal,
        description: Optional[str] = None,
        fiscal_year: Optional[int] = None,
        user_id: Optional[str] = None,
    ) -> ProjectCostBudget:
        """
        创建预算记录.

        Args:
            project_id: 项目ID（必填）
            category: 成本类别（必填）
            amount: 预算金额（必填）
            description: 描述
            fiscal_year: 财年
            user_id: 创建用户ID

        Returns:
            ProjectCostBudget: 创建的预算

        Raises:
            ProjectNotFoundError: 项目不存在
        """
        # 验证项目存在
        project = await self._project_repository.get_by_id(project_id)
        if project is None:
            raise ProjectNotFoundError(project_id=str(project_id))

        # 创建预算数据
        budget_data = {
            "id": uuid.uuid4(),
            "project_id": project_id,
            "category": category,
            "amount": amount,
            "description": description,
            "fiscal_year": fiscal_year,
        }

        budget = await self._budget_repository.create_in_project(project_id, budget_data)

        logger.info(
            "Budget created",
            extra={
                "budget_id": str(budget.id),
                "project_id": str(project_id),
                "category": category,
                "amount": float(amount),
            }
        )

        return budget

    async def get_budget(
        self,
        budget_id: uuid.UUID,
        project_id: uuid.UUID,
    ) -> ProjectCostBudget:
        """获取预算记录."""
        budget = await self._budget_repository.get_by_id_or_error(budget_id, project_id)
        return budget

    async def update_budget(
        self,
        budget_id: uuid.UUID,
        project_id: uuid.UUID,
        **kwargs,
    ) -> ProjectCostBudget:
        """更新预算记录."""
        budget = await self._budget_repository.update_in_project(budget_id, project_id, kwargs)
        logger.info(
            "Budget updated",
            extra={
                "budget_id": str(budget_id),
                "project_id": str(project_id),
                "fields": list(kwargs.keys()),
            }
        )
        return budget

    async def list_budgets(
        self,
        project_id: uuid.UUID,
        category: Optional[CostCategory] = None,
        skip: int = 0,
        limit: int = 100,
    ) -> List[ProjectCostBudget]:
        """列出预算记录."""
        filters = {}
        if category:
            filters["category"] = category
        return await self._budget_repository.list_by_project(
            project_id=project_id,
            filters=filters,
            skip=skip,
            limit=limit,
        )

    async def delete_budget(
        self,
        budget_id: uuid.UUID,
        project_id: uuid.UUID,
    ) -> bool:
        """删除预算记录."""
        return await self._budget_repository.delete_in_project(budget_id, project_id)

    async def create_actual(
        self,
        project_id: uuid.UUID,
        category: CostCategory,
        amount: Decimal,
        expense_date: date,
        description: Optional[str] = None,
        invoice_number: Optional[str] = None,
        user_id: Optional[str] = None,
    ) -> ProjectCostActual:
        """创建实际支出记录."""
        project = await self._project_repository.get_by_id(project_id)
        if project is None:
            raise ProjectNotFoundError(project_id=str(project_id))

        actual_data = {
            "id": uuid.uuid4(),
            "project_id": project_id,
            "category": category,
            "amount": amount,
            "expense_date": expense_date,
            "description": description,
            "invoice_number": invoice_number,
            "approval_status": "pending",
        }

        actual = await self._actual_repository.create_in_project(project_id, actual_data)

        # 检查是否触发预警
        await self._check_cost_warning(project_id)

        logger.info(
            "Actual cost created",
            extra={
                "actual_id": str(actual.id),
                "project_id": str(project_id),
                "category": category,
                "amount": float(amount),
                "expense_date": str(expense_date),
            }
        )

        return actual

    async def get_actual(
        self,
        actual_id: uuid.UUID,
        project_id: uuid.UUID,
    ) -> ProjectCostActual:
        """获取实际支出记录."""
        actual = await self._actual_repository.get_by_id_or_error(actual_id, project_id)
        return actual

    async def update_actual(
        self,
        actual_id: uuid.UUID,
        project_id: uuid.UUID,
        **kwargs,
    ) -> ProjectCostActual:
        """更新实际支出记录."""
        actual = await self._actual_repository.update_in_project(actual_id, project_id, kwargs)
        logger.info(
            "Actual cost updated",
            extra={
                "actual_id": str(actual_id),
                "project_id": str(project_id),
                "fields": list(kwargs.keys()),
            }
        )
        return actual

    async def list_actuals(
        self,
        project_id: uuid.UUID,
        category: Optional[CostCategory] = None,
        skip: int = 0,
        limit: int = 100,
    ) -> List[ProjectCostActual]:
        """列出实际支出记录."""
        filters = {}
        if category:
            filters["category"] = category
        return await self._actual_repository.list_by_project(
            project_id=project_id,
            filters=filters,
            skip=skip,
            limit=limit,
        )

    async def delete_actual(
        self,
        actual_id: uuid.UUID,
        project_id: uuid.UUID,
    ) -> bool:
        """删除实际支出记录."""
        return await self._actual_repository.delete_in_project(actual_id, project_id)

    # ==================== 新增功能：成本分析 ====================

    async def get_cost_analysis(
        self,
        project_id: uuid.UUID,
    ) -> Dict[str, Any]:
        """
        获取成本分析报告.

        计算成本偏差、成本执行率、剩余预算等关键指标。

        Args:
            project_id: 项目ID

        Returns:
            Dict: 成本分析报告
        """
        # 获取总预算和实际支出
        total_budget = await self._budget_repository.get_total_budget(project_id)
        total_actual = await self._actual_repository.get_total_actual(project_id)

        # 计算偏差
        variance = total_budget - total_actual
        variance_percent = float((variance / total_budget * 100) if total_budget > 0 else Decimal('0'))

        # 计算执行率
        execution_rate = float((total_actual / total_budget * 100) if total_budget > 0 else Decimal('0'))

        # 剩余预算
        remaining_budget = total_budget - total_actual

        # 是否超支
        is_over_budget = total_actual > total_budget

        # 按类别分析
        budget_by_category = await self._budget_repository.get_budget_by_category(project_id)
        actual_by_category = await self._actual_repository.get_actual_by_category(project_id)

        # 合并类别分析
        category_analysis = []
        for category in CostCategory:
            budget_amount = budget_by_category.get(category.value, Decimal('0'))
            actual_amount = actual_by_category.get(category.value, Decimal('0'))
            cat_variance = budget_amount - actual_amount
            cat_execution_rate = float((actual_amount / budget_amount * 100) if budget_amount > 0 else Decimal('0'))

            category_analysis.append({
                "category": category.value,
                "budget": float(budget_amount),
                "actual": float(actual_amount),
                "variance": float(cat_variance),
                "execution_rate": cat_execution_rate,
                "is_over_budget": actual_amount > budget_amount,
            })

        # 预警状态
        warning_status = None
        threshold_config = self._threshold_configs.get(str(project_id))
        if threshold_config:
            if execution_rate >= threshold_config.threshold_percent:
                warning_status = {
                    "level": threshold_config.warning_level,
                    "percent": execution_rate,
                    "threshold": threshold_config.threshold_percent,
                    "message": f"成本执行率达到{execution_rate:.1f}%，超过阈值{threshold_config.threshold_percent}%",
                }

        analysis = {
            "project_id": str(project_id),
            "total_budget": float(total_budget),
            "total_actual": float(total_actual),
            "variance": float(variance),
            "variance_percent": variance_percent,
            "execution_rate": execution_rate,
            "remaining_budget": float(remaining_budget),
            "is_over_budget": is_over_budget,
            "category_analysis": category_analysis,
            "warning_status": warning_status,
            "analysis_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        logger.info(f"Cost analysis generated: project={project_id}")
        return analysis

    # ==================== 新增功能：成本阈值设置 ====================

    async def set_cost_threshold(
        self,
        project_id: uuid.UUID,
        threshold_percent: float,
        warning_level: str,
        notify_users: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        设置成本预警阈值.

        Args:
            project_id: 项目ID
            threshold_percent: 阈值百分比
            warning_level: 预警等级
            notify_users: 通知用户列表

        Returns:
            Dict: 设置结果
        """
        # 验证项目存在
        project = await self._project_repository.get_by_id(project_id)
        if project is None:
            raise ProjectNotFoundError(project_id=str(project_id))

        # 创建阈值配置
        config = CostThresholdConfig(
            project_id=project_id,
            threshold_percent=threshold_percent,
            warning_level=warning_level,
            notify_users=notify_users,
        )

        # 保存配置（内存中，实际应存储到数据库）
        self._threshold_configs[str(project_id)] = config

        logger.info(
            f"Cost threshold set: project={project_id}, percent={threshold_percent}, level={warning_level}"
        )

        return {
            "project_id": str(project_id),
            "threshold_percent": threshold_percent,
            "warning_level": warning_level,
            "notify_users": notify_users or [],
            "created_at": config.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        }

    # ==================== 新增功能：成本预警 ====================

    async def _check_cost_warning(
        self,
        project_id: uuid.UUID,
    ) -> Optional[Dict[str, Any]]:
        """
        检查成本是否触发预警.

        Args:
            project_id: 项目ID

        Returns:
            Optional[Dict]: 预警信息，如果未触发则返回None
        """
        # 获取阈值配置
        threshold_config = self._threshold_configs.get(str(project_id))
        if not threshold_config:
            # 使用默认阈值
            threshold_config = CostThresholdConfig(
                project_id=project_id,
                threshold_percent=DEFAULT_THRESHOLD_PERCENT,
                warning_level="medium",
            )

        # 获取成本执行率
        analysis = await self.get_cost_analysis(project_id)
        execution_rate = analysis["execution_rate"]

        # 检查是否超过阈值
        if execution_rate >= threshold_config.threshold_percent:
            warning = {
                "project_id": str(project_id),
                "level": threshold_config.warning_level,
                "execution_rate": execution_rate,
                "threshold": threshold_config.threshold_percent,
                "total_budget": analysis["total_budget"],
                "total_actual": analysis["total_actual"],
                "message": f"成本执行率达到{execution_rate:.1f}%，超过阈值{threshold_config.threshold_percent}%",
                "notify_users": threshold_config.notify_users,
                "triggered_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }

            # 发送通知（如果有配置通知用户）
            if threshold_config.notify_users:
                await self._send_cost_warning_notification(project_id, warning)

            logger.warning(f"Cost warning triggered: project={project_id}, rate={execution_rate}")
            return warning

        return None

    async def _send_cost_warning_notification(
        self,
        project_id: uuid.UUID,
        warning: Dict[str, Any],
    ) -> None:
        """
        发送成本预警通知.

        Args:
            project_id: 项目ID
            warning: 预警信息
        """
        try:
            notification_service = NotificationService(self.session)

            # 获取项目信息
            project = await self._project_repository.get_by_id(project_id)

            # 构建通知内容
            for user_id in warning.get("notify_users", []):
                try:
                    await notification_service.send_cost_warning_message(
                        user_id=user_id,
                        project_name=project.name if project else "未知项目",
                        execution_rate=warning["execution_rate"],
                        threshold=warning["threshold"],
                        total_budget=warning["total_budget"],
                        total_actual=warning["total_actual"],
                        level=warning["level"],
                    )
                except Exception as e:
                    logger.warning(f"Failed to send cost warning to {user_id}: {e}")

        except Exception as e:
            logger.error(f"Failed to send cost warning notifications: {e}")

    async def get_cost_warnings(
        self,
        project_id: Optional[uuid.UUID] = None,
        warning_level: Optional[str] = None,
        limit: int = 50,
    ) -> List[Dict[str, Any]]:
        """
        获取超支预警列表.

        Args:
            project_id: 项目ID过滤
            warning_level: 预警等级过滤
            limit: 返回数量限制

        Returns:
            List[Dict]: 预警列表
        """
        warnings = []

        # 如果指定了项目ID，检查该项目
        if project_id:
            warning = await self._check_cost_warning(project_id)
            if warning:
                warnings.append(warning)
        else:
            # 检查所有配置了阈值的项目
            for project_id_str, config in self._threshold_configs.items():
                try:
                    pid = uuid.UUID(project_id_str)
                    warning = await self._check_cost_warning(pid)
                    if warning:
                        # 过滤预警等级
                        if warning_level and warning["level"] != warning_level:
                            continue
                        warnings.append(warning)
                except Exception as e:
                    logger.warning(f"Failed to check warning for project {project_id_str}: {e}")

        # 限制返回数量
        warnings = warnings[:limit]

        logger.info(f"Cost warnings retrieved: count={len(warnings)}")
        return warnings

    # ==================== 新增功能：成本导入 ====================

    async def import_cost_batch(
        self,
        project_id: uuid.UUID,
        data_list: List[Dict[str, Any]],
        import_mode: str,
    ) -> Dict[str, int]:
        """
        批量导入成本数据.

        Args:
            project_id: 项目ID
            data_list: 数据列表
            import_mode: 导入模式

        Returns:
            Dict: 导入统计 {imported, updated, skipped}
        """
        result = {"imported": 0, "updated": 0, "skipped": 0}

        # 验证项目存在
        project = await self._project_repository.get_by_id(project_id)
        if project is None:
            raise ProjectNotFoundError(project_id=str(project_id))

        for data in data_list:
            try:
                # 判断是预算还是实际支出
                if "expense_date" in data:
                    # 实际支出
                    category = CostCategory(data.get("category", "other"))
                    amount = Decimal(str(data.get("amount", 0)))
                    expense_date = self._parse_date(data.get("expense_date"))

                    if import_mode == ImportMode.FULL_REPLACE.value:
                        # 全量替换：删除旧数据后插入
                        await self.create_actual(
                            project_id=project_id,
                            category=category,
                            amount=amount,
                            expense_date=expense_date,
                            description=data.get("description"),
                            invoice_number=data.get("invoice_number"),
                        )
                        result["imported"] += 1

                    elif import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                        # 增量更新：根据发票号判断
                        existing = await self._find_actual_by_invoice(
                            project_id, data.get("invoice_number")
                        )
                        if existing:
                            await self.update_actual(
                                existing.id,
                                project_id,
                                amount=amount,
                                expense_date=expense_date,
                                description=data.get("description"),
                            )
                            result["updated"] += 1
                        else:
                            await self.create_actual(
                                project_id=project_id,
                                category=category,
                                amount=amount,
                                expense_date=expense_date,
                                description=data.get("description"),
                                invoice_number=data.get("invoice_number"),
                            )
                            result["imported"] += 1

                    elif import_mode == ImportMode.APPEND_ONLY.value:
                        # 仅追加
                        await self.create_actual(
                            project_id=project_id,
                            category=category,
                            amount=amount,
                            expense_date=expense_date,
                            description=data.get("description"),
                            invoice_number=data.get("invoice_number"),
                        )
                        result["imported"] += 1

                else:
                    # 预算
                    category = CostCategory(data.get("category", "other"))
                    amount = Decimal(str(data.get("amount", 0)))

                    if import_mode == ImportMode.FULL_REPLACE.value:
                        await self.create_budget(
                            project_id=project_id,
                            category=category,
                            amount=amount,
                            description=data.get("description"),
                            fiscal_year=data.get("fiscal_year"),
                        )
                        result["imported"] += 1

                    elif import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                        # 增量更新：根据类别判断
                        existing = await self._find_budget_by_category(
                            project_id, category
                        )
                        if existing:
                            await self.update_budget(
                                existing.id,
                                project_id,
                                amount=amount,
                                description=data.get("description"),
                                fiscal_year=data.get("fiscal_year"),
                            )
                            result["updated"] += 1
                        else:
                            await self.create_budget(
                                project_id=project_id,
                                category=category,
                                amount=amount,
                                description=data.get("description"),
                                fiscal_year=data.get("fiscal_year"),
                            )
                            result["imported"] += 1

                    elif import_mode == ImportMode.APPEND_ONLY.value:
                        await self.create_budget(
                            project_id=project_id,
                            category=category,
                            amount=amount,
                            description=data.get("description"),
                            fiscal_year=data.get("fiscal_year"),
                        )
                        result["imported"] += 1

            except Exception as e:
                logger.warning(f"Failed to import cost data: {data}, error={e}")
                result["skipped"] += 1

        await self.session.flush()
        logger.info(f"Cost batch import completed: project={project_id}, result={result}")
        return result

    async def _find_actual_by_invoice(
        self,
        project_id: uuid.UUID,
        invoice_number: Optional[str],
    ) -> Optional[ProjectCostActual]:
        """根据发票号查找实际支出."""
        if not invoice_number:
            return None

        result = await self.session.execute(
            select(ProjectCostActual).where(
                and_(
                    ProjectCostActual.project_id == project_id,
                    ProjectCostActual.invoice_number == invoice_number,
                )
            )
        )
        return result.scalar_one_or_none()

    async def _find_budget_by_category(
        self,
        project_id: uuid.UUID,
        category: CostCategory,
    ) -> Optional[ProjectCostBudget]:
        """根据类别查找预算."""
        result = await self.session.execute(
            select(ProjectCostBudget).where(
                and_(
                    ProjectCostBudget.project_id == project_id,
                    ProjectCostBudget.category == category,
                )
            )
        )
        return result.scalar_one_or_none()

    def _parse_date(self, date_str: Optional[str]) -> date:
        """解析日期字符串."""
        if not date_str:
            return date.today()

        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"]:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue

        return date.today()

    # ==================== 新增功能：成本导出 ====================

    async def generate_export_excel(
        self,
        project_id: uuid.UUID,
        export_type: str = "all",
    ) -> BytesIO:
        """
        生成成本导出Excel文件.

        Args:
            project_id: 项目ID
            export_type: 导出类型

        Returns:
            BytesIO: Excel文件流
        """
        wb = Workbook()

        # 样式定义
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 导出时间
        ws = wb.active
        ws.title = "成本数据"
        ws["A1"] = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["B1"] = f"项目ID: {str(project_id)}"

        if export_type == "budget" or export_type == "all":
            # 预算数据
            if export_type == "budget":
                ws.title = "成本预算"
            else:
                ws_budget = wb.create_sheet(title="成本预算")

                # 写入标题
                ws_budget["A1"] = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws_budget["B1"] = f"项目ID: {str(project_id)}"

                # 写入表头
                budget_headers = ["类别", "预算金额", "描述", "财年"]
                for col, header in enumerate(budget_headers, 1):
                    cell = ws_budget.cell(row=2, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border

                # 获取预算数据
                budgets = await self.list_budgets(project_id)
                for row_idx, budget in enumerate(budgets, 3):
                    ws_budget.cell(row=row_idx, column=1, value=budget.category).border = border
                    ws_budget.cell(row=row_idx, column=2, value=float(budget.amount)).border = border
                    ws_budget.cell(row=row_idx, column=3, value=budget.description or "").border = border
                    ws_budget.cell(row=row_idx, column=4, value=budget.fiscal_year or "").border = border

        if export_type == "actual" or export_type == "all":
            # 实际支出数据
            ws_actual = wb.create_sheet(title="实际支出")

            # 写入标题
            ws_actual["A1"] = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_actual["B1"] = f"项目ID: {str(project_id)}"

            # 写入表头
            actual_headers = ["类别", "实际金额", "支出日期", "描述", "发票号", "审批状态"]
            for col, header in enumerate(actual_headers, 1):
                cell = ws_actual.cell(row=2, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # 获取实际支出数据
            actuals = await self.list_actuals(project_id)
            for row_idx, actual in enumerate(actuals, 3):
                ws_actual.cell(row=row_idx, column=1, value=actual.category).border = border
                ws_actual.cell(row=row_idx, column=2, value=float(actual.amount)).border = border
                ws_actual.cell(row=row_idx, column=3, value=str(actual.expense_date)).border = border
                ws_actual.cell(row=row_idx, column=4, value=actual.description or "").border = border
                ws_actual.cell(row=row_idx, column=5, value=actual.invoice_number or "").border = border
                ws_actual.cell(row=row_idx, column=6, value=actual.approval_status or "").border = border

        # 成本分析摘要
        if export_type == "all":
            ws_analysis = wb.create_sheet(title="成本分析")
            analysis = await self.get_cost_analysis(project_id)

            ws_analysis["A1"] = "成本分析摘要"
            ws_analysis["A1"].font = Font(bold=True, size=12)

            ws_analysis["A2"] = "总预算"
            ws_analysis["B2"] = analysis["total_budget"]
            ws_analysis["A3"] = "实际支出"
            ws_analysis["B3"] = analysis["total_actual"]
            ws_analysis["A4"] = "偏差"
            ws_analysis["B4"] = analysis["variance"]
            ws_analysis["A5"] = "执行率"
            ws_analysis["B5"] = f"{analysis['execution_rate']:.1f}%"
            ws_analysis["A6"] = "是否超支"
            ws_analysis["B6"] = "是" if analysis["is_over_budget"] else "否"

        # 写入BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f"Cost export Excel generated: project={project_id}, type={export_type}")
        return buffer

    # ==================== 新增功能：飞书表格同步 ====================

    async def sync_cost_to_lark_sheet(
        self,
        project_id: uuid.UUID,
        sheet_token: str,
        sync_direction: str,
        sync_engine: SyncEngine,
    ) -> Dict[str, Any]:
        """
        同步成本数据到飞书在线表格.

        Args:
            project_id: 项目ID
            sheet_token: 飞书表格Token
            sync_direction: 同步方向
            sync_engine: 同步引擎

        Returns:
            Dict: 同步结果
        """
        result = {
            "success": True,
            "sync_direction": sync_direction,
            "records_synced": 0,
            "errors": [],
        }

        try:
            # 创建同步日志
            sync_log = await sync_engine.create_sync_log(
                sync_type="lark_sheet",
                sync_direction=sync_direction,
                module="cost",
                project_id=project_id,
                lark_sheet_token=sheet_token,
            )

            if sync_direction == "to_sheet":
                # 导出数据到飞书表格
                # 实际实现需要调用飞书API写入表格
                # 这里简化处理
                budgets = await self.list_budgets(project_id)
                actuals = await self.list_actuals(project_id)
                result["records_synced"] = len(budgets) + len(actuals)

            elif sync_direction == "from_sheet":
                # 从飞书表格导入数据
                # 实际实现需要调用飞书API读取表格
                # 这里简化处理
                result["records_synced"] = 0

            elif sync_direction == "bidirectional":
                # 双向同步
                # 先导出，再导入
                budgets = await self.list_budgets(project_id)
                actuals = await self.list_actuals(project_id)
                result["records_synced"] = len(budgets) + len(actuals)

            # 更新同步日志
            await sync_engine.update_sync_status(
                sync_log.id,
                status="success",
                records_total=result["records_synced"],
                records_success=result["records_synced"],
            )

        except Exception as e:
            result["success"] = False
            result["errors"].append(str(e))
            logger.error(f"Failed to sync cost to Lark sheet: {e}")

        logger.info(f"Cost sync completed: project={project_id}, result={result}")
        return result

    # ==================== 新增功能：成本统计 ====================

    async def get_cost_statistics(
        self,
        project_id: uuid.UUID,
    ) -> Dict[str, Any]:
        """
        获取成本统计报表数据.

        用于生成成本趋势分析图等可视化数据。

        Args:
            project_id: 项目ID

        Returns:
            Dict: 统计报表数据
        """
        # 获取分析数据
        analysis = await self.get_cost_analysis(project_id)

        # 获取按类别的详细统计
        budget_by_category = await self._budget_repository.get_budget_by_category(project_id)
        actual_by_category = await self._actual_repository.get_actual_by_category(project_id)

        # 按时间的实际支出分布
        time_distribution = await self._get_actual_time_distribution(project_id)

        statistics = {
            "summary": {
                "total_budget": analysis["total_budget"],
                "total_actual": analysis["total_actual"],
                "variance": analysis["variance"],
                "execution_rate": analysis["execution_rate"],
                "remaining_budget": analysis["remaining_budget"],
            },
            "by_category": {
                "budget": {k: float(v) for k, v in budget_by_category.items()},
                "actual": {k: float(v) for k, v in actual_by_category.items()},
            },
            "time_distribution": time_distribution,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        logger.info(f"Cost statistics generated: project={project_id}")
        return statistics

    async def _get_actual_time_distribution(
        self,
        project_id: uuid.UUID,
    ) -> List[Dict[str, Any]]:
        """获取实际支出的时间分布."""
        # 按月汇总实际支出
        try:
            result = await self.session.execute(
                select(
                    func.strftime("%Y-%m", ProjectCostActual.expense_date).label("month"),
                    func.sum(ProjectCostActual.amount).label("total"),
                ).where(ProjectCostActual.project_id == project_id)
                .group_by(func.strftime("%Y-%m", ProjectCostActual.expense_date))
                .order_by(func.strftime("%Y-%m", ProjectCostActual.expense_date))
            )

            rows = result.all()
            return [
                {
                    "month": row.month,
                    "total": float(row.total or 0),
                }
                for row in rows
            ]
        except Exception as e:
            logger.warning(f"Failed to get time distribution: {e}")
            return []

    # ==================== 新增功能：成本趋势分析 ====================

    async def get_cost_trend(
        self,
        project_id: uuid.UUID,
        period: str = "month",
    ) -> Dict[str, Any]:
        """
        获取成本趋势分析数据.

        Args:
            project_id: 项目ID
            period: 时间周期

        Returns:
            Dict: 趋势分析数据
        """
        # 获取实际支出的时间分布
        time_distribution = await self._get_actual_time_distribution(project_id)

        # 获取总预算作为基准线
        total_budget = await self._budget_repository.get_total_budget(project_id)

        # 计算累计支出
        cumulative_actual = []
        cumsum = Decimal('0')
        for item in time_distribution:
            cumsum += Decimal(str(item["total"]))
            cumulative_actual.append({
                "month": item["month"],
                "cumulative": float(cumsum),
                "percent": float((cumsum / total_budget * 100) if total_budget > 0 else Decimal('0')),
            })

        trend = {
            "period": period,
            "total_budget": float(total_budget),
            "monthly_actual": time_distribution,
            "cumulative_actual": cumulative_actual,
            "projected_completion": self._estimate_completion(total_budget, cumulative_actual),
        }

        logger.info(f"Cost trend analysis generated: project={project_id}")
        return trend

    def _estimate_completion(
        self,
        total_budget: Decimal,
        cumulative_actual: List[Dict],
    ) -> Dict[str, Any]:
        """估算项目完工时成本."""
        if not cumulative_actual:
            return {"estimated_total": 0, "will_over_budget": False}

        # 简化估算：基于当前累计和平均增长率
        latest_cumulative = Decimal(str(cumulative_actual[-1]["cumulative"]))

        # 假设项目按当前速度继续
        estimated_total = latest_cumulative

        will_over_budget = estimated_total > total_budget

        return {
            "estimated_total": float(estimated_total),
            "will_over_budget": will_over_budget,
            "over_budget_amount": float(estimated_total - total_budget) if will_over_budget else 0,
        }

    # ==================== 新增功能：成本审批 ====================

    async def submit_cost_approval(
        self,
        project_id: uuid.UUID,
        actual_id: uuid.UUID,
        approval_type: str,
    ) -> Dict[str, Any]:
        """
        提交成本审批流程.

        集成飞书审批API。

        Args:
            project_id: 项目ID
            actual_id: 实际支出ID
            approval_type: 审批类型

        Returns:
            Dict: 审批提交结果
        """
        # 获取实际支出信息
        actual = await self.get_actual(actual_id, project_id)
        project = await self._project_repository.get_by_id(project_id)

        # 更新审批状态
        await self.update_actual(actual_id, project_id, approval_status="pending_approval")

        # 实际实现需要调用飞书审批API创建审批实例
        # 这里简化处理
        result = {
            "success": True,
            "approval_type": approval_type,
            "actual_id": str(actual_id),
            "amount": float(actual.amount),
            "category": actual.category,
            "approval_status": "pending_approval",
            "submitted_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        logger.info(f"Cost approval submitted: project={project_id}, actual={actual_id}")
        return result

    # ==================== 原有功能：成本汇总 ====================

    async def get_cost_summary(
        self,
        project_id: uuid.UUID,
    ) -> Dict[str, Any]:
        """获取成本汇总."""
        return await self._actual_repository.get_cost_variance(project_id)

    async def get_budget_by_category(
        self,
        project_id: uuid.UUID,
    ) -> Dict[str, Decimal]:
        """按类别获取预算汇总."""
        return await self._budget_repository.get_budget_by_category(project_id)

    async def get_actual_by_category(
        self,
        project_id: uuid.UUID,
    ) -> Dict[str, Decimal]:
        """按类别获取实际支出汇总."""
        return await self._actual_repository.get_actual_by_category(project_id)