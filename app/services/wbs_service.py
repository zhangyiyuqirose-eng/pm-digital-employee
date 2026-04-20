"""
PM Digital Employee - WBS Service
项目经理数字员工系统 - WBS分解业务服务

v1.2.0新增：
- 导入WBS树形结构（Excel、飞书表格）
- 导出WBS为Excel（树形结构格式）
- WBS分解结果自动同步到Task模块（创建任务）
- WBS版本管理（创建新版本、发布、回滚）
- WBS可视化数据生成（树形结构、甘特图数据）
"""

import json
import uuid
from datetime import datetime, timezone, date, timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import select, and_, or_, func, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import ErrorCode, ProjectNotFoundError, TaskNotFoundError
from app.core.logging import get_logger
from app.domain.enums import TaskStatus, TaskPriority, WBSStatus, DataSource, DependencyType
from app.domain.models.wbs_version import WBSVersion
from app.domain.models.task import Task
from app.domain.models.project import Project
from app.repositories.project_repository import ProjectRepository
from app.services.task_service import TaskService
from app.services.excel_service import ExcelService
from app.services.sync_engine import SyncEngine

logger = get_logger(__name__)


class WBSNode:
    """
    WBS节点数据结构.

    用于表示WBS树形结构中的单个节点。

    Attributes:
        id: 节点ID
        name: 节点名称
        level: 层级（1=项目级，2=任务级，3=子任务级...）
        parent_id: 父节点ID
        duration: 工期天数
        start_date: 开始日期
        end_date: 结束日期
        assignee_id: 负责人ID
        assignee_name: 负责人姓名
        deliverable: 交付物
        dependencies: 前置依赖节点ID列表
        children: 子节点列表
    """

    def __init__(
        self,
        id: str,
        name: str,
        level: int,
        parent_id: Optional[str] = None,
        duration: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        assignee_id: Optional[str] = None,
        assignee_name: Optional[str] = None,
        deliverable: Optional[str] = None,
        dependencies: Optional[List[str]] = None,
        description: Optional[str] = None,
    ) -> None:
        self.id = id
        self.name = name
        self.level = level
        self.parent_id = parent_id
        self.duration = duration
        self.start_date = start_date
        self.end_date = end_date
        self.assignee_id = assignee_id
        self.assignee_name = assignee_name
        self.deliverable = deliverable
        self.dependencies = dependencies or []
        self.description = description
        self.children: List[WBSNode] = []

    def to_dict(self) -> Dict[str, Any]:
        """
        转换为字典格式.

        Returns:
            Dict: 节点数据字典
        """
        return {
            "id": self.id,
            "name": self.name,
            "level": self.level,
            "parent_id": self.parent_id,
            "duration": self.duration,
            "start_date": self.start_date.isoformat() if self.start_date else None,
            "end_date": self.end_date.isoformat() if self.end_date else None,
            "assignee_id": self.assignee_id,
            "assignee_name": self.assignee_name,
            "deliverable": self.deliverable,
            "dependencies": self.dependencies,
            "description": self.description,
            "children": [child.to_dict() for child in self.children],
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "WBSNode":
        """
        从字典创建节点.

        Args:
            data: 节点数据字典

        Returns:
            WBSNode: 创建的节点对象
        """
        # 处理日期字段
        start_date = None
        if data.get("start_date"):
            if isinstance(data["start_date"], str):
                start_date = date.fromisoformat(data["start_date"])
            elif isinstance(data["start_date"], date):
                start_date = data["start_date"]

        end_date = None
        if data.get("end_date"):
            if isinstance(data["end_date"], str):
                end_date = date.fromisoformat(data["end_date"])
            elif isinstance(data["end_date"], date):
                end_date = data["end_date"]

        node = cls(
            id=data.get("id", ""),
            name=data.get("name", ""),
            level=data.get("level", 1),
            parent_id=data.get("parent_id"),
            duration=data.get("duration"),
            start_date=start_date,
            end_date=end_date,
            assignee_id=data.get("assignee_id"),
            assignee_name=data.get("assignee_name"),
            deliverable=data.get("deliverable"),
            dependencies=data.get("dependencies", []),
            description=data.get("description"),
        )

        # 递归创建子节点
        for child_data in data.get("children", []):
            node.children.append(cls.from_dict(child_data))

        return node


class WBSService:
    """
    WBS分解业务服务.

    封装WBS相关的业务逻辑，包括：
    - WBS树形结构导入导出
    - WBS版本管理
    - WBS到任务的自动同步
    - WBS可视化数据生成
    """

    def __init__(self, session: AsyncSession) -> None:
        """
        初始化WBS服务.

        Args:
            session: 数据库会话
        """
        self.session = session
        self._project_repository = ProjectRepository(session)
        self._task_service = TaskService(session)
        self._excel_service = ExcelService(session)
        self._sync_engine = SyncEngine(session)

    # ==================== WBS导入导出 ====================

    async def import_wbs_from_excel(
        self,
        file_path: str,
        project_id: uuid.UUID,
        version_name: Optional[str] = None,
        description: Optional[str] = None,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
    ) -> WBSVersion:
        """
        从Excel导入WBS树形结构.

        Excel模板格式：
        - 第1列：层级编号（1, 2, 3...）
        - 第2列：节点名称
        - 第3列：工期（天数）
        - 第4列：开始日期
        - 第5列：结束日期
        - 第6列：负责人
        - 第7列：交付物
        - 第8列：前置任务（节点ID）

        Args:
            file_path: Excel文件路径
            project_id: 项目ID
            version_name: 版本名称
            description: 版本描述
            operator_id: 操作人ID
            operator_name: 操作人姓名

        Returns:
            WBSVersion: 创建的WBS版本

        Raises:
            ProjectNotFoundError: 项目不存在
            ValueError: Excel解析失败
        """
        # 验证项目存在
        project = await self._project_repository.get_by_id(project_id)
        if project is None:
            raise ProjectNotFoundError(project_id=str(project_id))

        # 解析Excel文件
        wbs_tree = self._parse_wbs_excel(file_path)

        # 创建新版本
        version = await self.create_version(
            project_id=project_id,
            wbs_data=wbs_tree,
            version_name=version_name,
            description=description,
            data_source=DataSource.EXCEL_IMPORT,
            created_by_id=operator_id,
            created_by_name=operator_name,
        )

        logger.info(
            "WBS imported from Excel",
            extra={
                "version_id": str(version.id),
                "project_id": str(project_id),
                "file_path": file_path,
                "nodes_count": self._count_nodes(wbs_tree),
            }
        )

        return version

    async def import_wbs_from_lark_sheet(
        self,
        sheet_token: str,
        sheet_id: str,
        project_id: uuid.UUID,
        version_name: Optional[str] = None,
        description: Optional[str] = None,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
    ) -> WBSVersion:
        """
        从飞书表格导入WBS树形结构.

        Args:
            sheet_token: 飞书表格Token
            sheet_id: 子表格ID
            project_id: 项目ID
            version_name: 版本名称
            description: 版本描述
            operator_id: 操作人ID
            operator_name: 操作人姓名

        Returns:
            WBSVersion: 创建的WBS版本
        """
        # 验证项目存在
        project = await self._project_repository.get_by_id(project_id)
        if project is None:
            raise ProjectNotFoundError(project_id=str(project_id))

        # 获取飞书表格数据（使用LarkSheetSyncService的逻辑）
        from app.services.lark_sheet_sync_service import LarkSheetSyncService
        sync_service = LarkSheetSyncService(self.session)
        
        sheet_data = await sync_service.fetch_sheet_data(sheet_token, sheet_id)
        
        # 解析表格数据为WBS树
        wbs_tree = self._parse_wbs_sheet_data(sheet_data)

        # 创建新版本
        version = await self.create_version(
            project_id=project_id,
            wbs_data=wbs_tree,
            version_name=version_name,
            description=description,
            data_source=DataSource.LARK_SHEET_SYNC,
            created_by_id=operator_id,
            created_by_name=operator_name,
        )

        logger.info(
            "WBS imported from Lark sheet",
            extra={
                "version_id": str(version.id),
                "project_id": str(project_id),
                "sheet_token": sheet_token,
            }
        )

        return version

    async def export_wbs_to_excel(
        self,
        project_id: uuid.UUID,
        version_id: Optional[uuid.UUID] = None,
    ) -> BytesIO:
        """
        导出WBS为Excel（树形结构格式）.

        Args:
            project_id: 项目ID
            version_id: 版本ID（可选，默认使用当前版本）

        Returns:
            BytesIO: Excel文件流
        """
        # 获取WBS版本
        if version_id:
            version = await self.get_version_by_id(version_id, project_id)
        else:
            version = await self.get_current_version(project_id)

        if not version:
            # 没有WBS版本，返回空模板
            return self._excel_service.generate_template("wbs")

        # 解析WBS数据
        wbs_tree = json.loads(version.wbs_data)

        # 生成Excel文件
        buffer = self._generate_wbs_excel(wbs_tree, project_id)

        logger.info(
            "WBS exported to Excel",
            extra={
                "project_id": str(project_id),
                "version_id": str(version.id) if version else None,
            }
        )

        return buffer

    # ==================== WBS版本管理 ====================

    async def create_version(
        self,
        project_id: uuid.UUID,
        wbs_data: Dict[str, Any],
        version_name: Optional[str] = None,
        description: Optional[str] = None,
        data_source: DataSource = DataSource.LARK_CARD,
        created_by_id: Optional[str] = None,
        created_by_name: Optional[str] = None,
    ) -> WBSVersion:
        """
        创建新的WBS版本.

        Args:
            project_id: 项目ID
            wbs_data: WBS树形结构数据
            version_name: 版本名称
            description: 版本描述
            data_source: 数据来源
            created_by_id: 创建者ID
            created_by_name: 创建者姓名

        Returns:
            WBSVersion: 创建的WBS版本

        Raises:
            ProjectNotFoundError: 项目不存在
        """
        # 验证项目存在
        project = await self._project_repository.get_by_id(project_id)
        if project is None:
            raise ProjectNotFoundError(project_id=str(project_id))

        # 获取当前最大版本号
        result = await self.session.execute(
            select(func.max(WBSVersion.version_number)).where(
                WBSVersion.project_id == project_id
            )
        )
        max_version = result.scalar_one_or_none() or 0
        new_version_number = max_version + 1

        # 生成版本名称
        if not version_name:
            version_name = f"V{new_version_number}"

        # 序列化WBS数据
        wbs_data_json = json.dumps(wbs_data, ensure_ascii=False)

        # 创建版本记录
        version = WBSVersion(
            id=uuid.uuid4(),
            project_id=project_id,
            version_number=new_version_number,
            version_name=version_name,
            description=description,
            wbs_data=wbs_data_json,
            status=WBSStatus.DRAFT,
            is_published=False,
            is_current=False,  # 新版本默认不是当前版本
            created_by_id=created_by_id,
            created_by_name=created_by_name,
            data_source=data_source,
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc),
        )

        self.session.add(version)
        await self.session.flush()
        await self.session.refresh(version)

        logger.info(
            "WBS version created",
            extra={
                "version_id": str(version.id),
                "project_id": str(project_id),
                "version_number": new_version_number,
                "data_source": data_source,
            }
        )

        return version

    async def publish_version(
        self,
        version_id: uuid.UUID,
        project_id: uuid.UUID,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
        sync_to_tasks: bool = True,
    ) -> WBSVersion:
        """
        发布WBS版本.

        发布后该版本成为当前版本，可选自动同步到任务模块。

        Args:
            version_id: 版本ID
            project_id: 项目ID（用于权限检查）
            operator_id: 发布者ID
            operator_name: 发布者姓名
            sync_to_tasks: 是否同步到任务模块

        Returns:
            WBSVersion: 发布后的版本

        Raises:
            ValueError: 版本不存在或已发布
        """
        # 获取版本
        version = await self.get_version_by_id(version_id, project_id)

        if not version:
            raise ValueError(f"WBS版本不存在: {version_id}")

        if version.is_published:
            raise ValueError(f"WBS版本已发布: {version_id}")

        # 将其他版本设为非当前版本
        await self.session.execute(
            update(WBSVersion)
            .where(WBSVersion.project_id == project_id)
            .values(is_current=False)
        )

        # 更新版本状态
        version.is_published = True
        version.is_current = True
        version.status = WBSStatus.PUBLISHED
        version.published_at = datetime.now(timezone.utc)
        version.published_by_id = operator_id
        version.published_by_name = operator_name
        version.updated_at = datetime.now(timezone.utc)

        await self.session.flush()
        await self.session.refresh(version)

        # 如果需要，同步到任务模块
        if sync_to_tasks:
            await self.wbs_to_tasks(version_id, project_id)

        logger.info(
            "WBS version published",
            extra={
                "version_id": str(version_id),
                "project_id": str(project_id),
                "sync_to_tasks": sync_to_tasks,
            }
        )

        return version

    async def rollback_to_version(
        self,
        version_id: uuid.UUID,
        project_id: uuid.UUID,
        operator_id: Optional[str] = None,
        operator_name: Optional[str] = None,
        sync_to_tasks: bool = True,
    ) -> WBSVersion:
        """
        回滚到指定WBS版本.

        Args:
            version_id: 要回滚到的版本ID
            project_id: 项目ID
            operator_id: 操作者ID
            operator_name: 操作者姓名
            sync_to_tasks: 是否同步到任务模块

        Returns:
            WBSVersion: 回滚后的当前版本

        Raises:
            ValueError: 版本不存在或未发布
        """
        # 获取版本
        version = await self.get_version_by_id(version_id, project_id)

        if not version:
            raise ValueError(f"WBS版本不存在: {version_id}")

        if not version.is_published:
            raise ValueError(f"只能回滚到已发布的版本: {version_id}")

        # 将其他版本设为非当前版本
        await self.session.execute(
            update(WBSVersion)
            .where(WBSVersion.project_id == project_id)
            .values(is_current=False)
        )

        # 设置指定版本为当前版本
        version.is_current = True
        version.updated_at = datetime.now(timezone.utc)

        await self.session.flush()
        await self.session.refresh(version)

        # 如果需要，同步到任务模块
        if sync_to_tasks:
            await self.wbs_to_tasks(version_id, project_id, clear_existing=True)

        logger.info(
            "WBS version rolled back",
            extra={
                "version_id": str(version_id),
                "project_id": str(project_id),
                "sync_to_tasks": sync_to_tasks,
            }
        )

        return version

    async def get_version_history(
        self,
        project_id: uuid.UUID,
        include_draft: bool = True,
        limit: int = 50,
    ) -> List[WBSVersion]:
        """
        获取WBS版本历史.

        Args:
            project_id: 项目ID
            include_draft: 是否包含草稿版本
            limit: 返回数量限制

        Returns:
            List[WBSVersion]: 版本列表（按版本号倒序）
        """
        query = select(WBSVersion).where(
            WBSVersion.project_id == project_id
        ).order_by(WBSVersion.version_number.desc())

        if not include_draft:
            query = query.where(WBSVersion.is_published == True)

        query = query.limit(limit)

        result = await self.session.execute(query)
        return list(result.scalars().all())

    async def get_version_by_id(
        self,
        version_id: uuid.UUID,
        project_id: uuid.UUID,
    ) -> Optional[WBSVersion]:
        """
        根据ID获取WBS版本.

        Args:
            version_id: 版本ID
            project_id: 项目ID（用于权限检查）

        Returns:
            Optional[WBSVersion]: 版本对象或None
        """
        result = await self.session.execute(
            select(WBSVersion).where(
                and_(
                    WBSVersion.id == version_id,
                    WBSVersion.project_id == project_id,
                )
            )
        )
        return result.scalar_one_or_none()

    async def get_current_version(
        self,
        project_id: uuid.UUID,
    ) -> Optional[WBSVersion]:
        """
        获取当前WBS版本.

        Args:
            project_id: 项目ID

        Returns:
            Optional[WBSVersion]: 当前版本或None
        """
        result = await self.session.execute(
            select(WBSVersion).where(
                and_(
                    WBSVersion.project_id == project_id,
                    WBSVersion.is_current == True,
                )
            ).order_by(WBSVersion.version_number.desc())
        )
        return result.scalar_one_or_none()

    # ==================== WBS到任务同步 ====================

    async def wbs_to_tasks(
        self,
        version_id: uuid.UUID,
        project_id: uuid.UUID,
        clear_existing: bool = False,
        operator_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        将WBS节点转换为任务.

        Args:
            version_id: WBS版本ID
            project_id: 项目ID
            clear_existing: 是否清除现有任务（用于回滚场景）
            operator_id: 操作者ID

        Returns:
            Dict: 同步结果统计
        """
        # 获取WBS版本
        version = await self.get_version_by_id(version_id, project_id)
        if not version:
            raise ValueError(f"WBS版本不存在: {version_id}")

        # 解析WBS数据
        wbs_data = json.loads(version.wbs_data)

        # 如果需要清除现有任务
        if clear_existing:
            await self._clear_wbs_tasks(project_id)

        # 创建任务映射表（WBS节点ID -> 任务ID）
        task_mapping: Dict[str, uuid.UUID] = {}

        # 转换WBS节点为任务
        stats = {
            "total_nodes": 0,
            "tasks_created": 0,
            "tasks_updated": 0,
            "errors": [],
        }

        # 遍历WBS树创建任务
        await self._convert_wbs_nodes_to_tasks(
            project_id=project_id,
            wbs_data=wbs_data,
            task_mapping=task_mapping,
            stats=stats,
            operator_id=operator_id,
        )

        # 设置任务依赖关系
        await self._set_task_dependencies(
            project_id=project_id,
            wbs_data=wbs_data,
            task_mapping=task_mapping,
        )

        logger.info(
            "WBS synced to tasks",
            extra={
                "version_id": str(version_id),
                "project_id": str(project_id),
                "tasks_created": stats["tasks_created"],
            }
        )

        return stats

    async def _convert_wbs_nodes_to_tasks(
        self,
        project_id: uuid.UUID,
        wbs_data: Dict[str, Any],
        task_mapping: Dict[str, uuid.UUID],
        stats: Dict[str, Any],
        operator_id: Optional[str] = None,
        parent_task_id: Optional[uuid.UUID] = None,
    ) -> None:
        """
        递归转换WBS节点为任务.

        Args:
            project_id: 项目ID
            wbs_data: WBS节点数据
            task_mapping: 节点ID到任务ID的映射
            stats: 统计信息
            operator_id: 操作者ID
            parent_task_id: 父任务ID
        """
        node = WBSNode.from_dict(wbs_data)
        stats["total_nodes"] += 1

        # 第1层级（项目级）不创建任务，只处理子节点
        if node.level == 1:
            for child_data in wbs_data.get("children", []):
                await self._convert_wbs_nodes_to_tasks(
                    project_id=project_id,
                    wbs_data=child_data,
                    task_mapping=task_mapping,
                    stats=stats,
                    operator_id=operator_id,
                    parent_task_id=None,
                )
            return

        # 创建任务
        try:
            # 检查是否已存在同名任务（根据WBS节点ID）
            existing_task = await self._find_task_by_wbs_node_id(
                project_id, node.id
            )

            if existing_task:
                # 更现有任务
                task = await self._task_service.update_task(
                    task_id=existing_task.id,
                    project_id=project_id,
                    name=node.name,
                    description=node.description,
                    start_date=node.start_date,
                    end_date=node.end_date,
                    estimated_duration=node.duration,
                    assignee_id=node.assignee_id,
                    assignee_name=node.assignee_name,
                    deliverable=node.deliverable,
                    parent_task_id=parent_task_id,
                )
                stats["tasks_updated"] += 1
            else:
                # 创建新任务
                task = await self._task_service.create_task(
                    project_id=project_id,
                    name=node.name,
                    description=node.description,
                    start_date=node.start_date,
                    end_date=node.end_date,
                    estimated_duration=node.duration,
                    assignee_id=node.assignee_id,
                    assignee_name=node.assignee_name,
                    deliverable=node.deliverable,
                    parent_task_id=parent_task_id,
                    priority=TaskPriority.MEDIUM,
                    status=TaskStatus.PENDING,
                    data_source=DataSource.SYSTEM_GENERATED,
                    user_id=operator_id,
                )
                stats["tasks_created"] += 1

            # 记录映射关系
            task_mapping[node.id] = task.id

            # 递归处理子节点
            for child_data in wbs_data.get("children", []):
                await self._convert_wbs_nodes_to_tasks(
                    project_id=project_id,
                    wbs_data=child_data,
                    task_mapping=task_mapping,
                    stats=stats,
                    operator_id=operator_id,
                    parent_task_id=task.id,
                )

        except Exception as e:
            stats["errors"].append({
                "node_id": node.id,
                "node_name": node.name,
                "error": str(e),
            })
            logger.warning(
                f"Failed to convert WBS node to task: {node.id} - {e}"
            )

    async def _set_task_dependencies(
        self,
        project_id: uuid.UUID,
        wbs_data: Dict[str, Any],
        task_mapping: Dict[str, uuid.UUID],
    ) -> None:
        """
        设置任务依赖关系.

        Args:
            project_id: 项目ID
            wbs_data: WBS数据
            task_mapping: 节点ID到任务ID的映射
        """
        node = WBSNode.from_dict(wbs_data)

        # 设置当前节点的依赖关系
        if node.dependencies and node.id in task_mapping:
            current_task_id = task_mapping[node.id]

            for dep_node_id in node.dependencies:
                if dep_node_id in task_mapping:
                    predecessor_task_id = task_mapping[dep_node_id]

                    try:
                        await self._task_service.set_task_dependency(
                            task_id=current_task_id,
                            project_id=project_id,
                            predecessor_task_id=predecessor_task_id,
                            dependency_type=DependencyType.FS,
                            adjust_dates=False,  # 不自动调整日期，保持WBS原始日期
                        )
                    except ValueError as e:
                        logger.warning(
                            f"Failed to set task dependency: {current_task_id} -> {predecessor_task_id}: {e}"
                        )

        # 递归处理子节点
        for child_data in wbs_data.get("children", []):
            await self._set_task_dependencies(
                project_id=project_id,
                wbs_data=child_data,
                task_mapping=task_mapping,
            )

    async def _find_task_by_wbs_node_id(
        self,
        project_id: uuid.UUID,
        wbs_node_id: str,
    ) -> Optional[Task]:
        """
        根据WBS节点ID查找任务.

        通过任务编码或名称匹配。

        Args:
            project_id: 项目ID
            wbs_node_id: WBS节点ID

        Returns:
            Optional[Task]: 匹配的任务或None
        """
        # 暂时通过名称匹配，后续可以添加WBS节点ID字段到Task模型
        result = await self.session.execute(
            select(Task).where(
                and_(
                    Task.project_id == project_id,
                    Task.code == wbs_node_id,
                )
            )
        )
        return result.scalar_one_or_none()

    async def _clear_wbs_tasks(
        self,
        project_id: uuid.UUID,
    ) -> None:
        """
        清除项目中的WBS生成任务.

        Args:
            project_id: 项目ID
        """
        # 删除所有系统生成的任务
        result = await self.session.execute(
            select(Task).where(
                and_(
                    Task.project_id == project_id,
                    Task.data_source == DataSource.SYSTEM_GENERATED,
                )
            )
        )
        tasks = result.scalars().all()

        for task in tasks:
            await self.session.delete(task)

        await self.session.flush()

        logger.info(
            f"Cleared {len(tasks)} WBS-generated tasks for project {project_id}"
        )

    # ==================== WBS可视化数据生成 ====================

    async def get_wbs_tree(
        self,
        project_id: uuid.UUID,
        version_id: Optional[uuid.UUID] = None,
    ) -> Dict[str, Any]:
        """
        获取WBS树形结构可视化数据.

        Args:
            project_id: 项目ID
            version_id: 版本ID（可选，默认使用当前版本）

        Returns:
            Dict: WBS树形结构数据
        """
        # 获取WBS版本
        if version_id:
            version = await self.get_version_by_id(version_id, project_id)
        else:
            version = await self.get_current_version(project_id)

        if not version:
            return {
                "project_id": str(project_id),
                "version_id": None,
                "tree": None,
                "message": "项目暂无WBS数据",
            }

        # 解析WBS数据
        wbs_tree = json.loads(version.wbs_data)

        return {
            "project_id": str(project_id),
            "version_id": str(version.id),
            "version_name": version.version_name,
            "version_number": version.version_number,
            "is_current": version.is_current,
            "tree": wbs_tree,
        }

    async def get_gantt_data(
        self,
        project_id: uuid.UUID,
        version_id: Optional[uuid.UUID] = None,
    ) -> Dict[str, Any]:
        """
        获取甘特图可视化数据.

        Args:
            project_id: 项目ID
            version_id: 版本ID（可选，默认使用当前版本）

        Returns:
            Dict: 甘特图数据（包含任务列表、依赖关系、时间线）
        """
        # 获取WBS数据
        wbs_tree_data = await self.get_wbs_tree(project_id, version_id)

        if not wbs_tree_data.get("tree"):
            return {
                "project_id": str(project_id),
                "tasks": [],
                "dependencies": [],
                "timeline": {
                    "start": None,
                    "end": None,
                },
            }

        # 解析WBS树并生成甘特图数据
        wbs_tree = wbs_tree_data["tree"]
        tasks: List[Dict[str, Any]] = []
        dependencies: List[Dict[str, Any]] = []

        # 扁平化WBS树并提取甘特图任务
        self._flatten_wbs_for_gantt(
            wbs_tree, tasks, dependencies, parent_id=None
        )

        # 计算时间线范围
        timeline_start = None
        timeline_end = None

        for task in tasks:
            if task.get("start"):
                task_start = date.fromisoformat(task["start"])
                if timeline_start is None or task_start < timeline_start:
                    timeline_start = task_start

            if task.get("end"):
                task_end = date.fromisoformat(task["end"])
                if timeline_end is None or task_end > timeline_end:
                    timeline_end = task_end

        return {
            "project_id": str(project_id),
            "version_id": wbs_tree_data.get("version_id"),
            "version_name": wbs_tree_data.get("version_name"),
            "tasks": tasks,
            "dependencies": dependencies,
            "timeline": {
                "start": timeline_start.isoformat() if timeline_start else None,
                "end": timeline_end.isoformat() if timeline_end else None,
            },
        }

    def _flatten_wbs_for_gantt(
        self,
        wbs_data: Dict[str, Any],
        tasks: List[Dict[str, Any]],
        dependencies: List[Dict[str, Any]],
        parent_id: Optional[str] = None,
    ) -> None:
        """
        扁平化WBS树为甘特图任务列表.

        Args:
            wbs_data: WBS节点数据
            tasks: 任务列表
            dependencies: 依赖关系列表
            parent_id: 父节点ID
        """
        node = WBSNode.from_dict(wbs_data)

        # 第1层级（项目级）不添加到甘特图
        if node.level > 1:
            gantt_task = {
                "id": node.id,
                "name": node.name,
                "level": node.level,
                "parent_id": parent_id,
                "start": node.start_date.isoformat() if node.start_date else None,
                "end": node.end_date.isoformat() if node.end_date else None,
                "duration": node.duration,
                "assignee": node.assignee_name,
                "assignee_id": node.assignee_id,
                "deliverable": node.deliverable,
                "progress": 0,  # WBS本身不含进度信息
            }
            tasks.append(gantt_task)

            # 添加依赖关系
            for dep_id in node.dependencies:
                dependencies.append({
                    "from": dep_id,
                    "to": node.id,
                    "type": "FS",  # 默认完成-开始依赖
                })

        # 递归处理子节点
        for child_data in wbs_data.get("children", []):
            self._flatten_wbs_for_gantt(
                child_data,
                tasks,
                dependencies,
                parent_id=node.id if node.level > 1 else None,
            )

    # ==================== Excel解析和生成辅助方法 ====================

    def _parse_wbs_excel(
        self,
        file_path: str,
    ) -> Dict[str, Any]:
        """
        解析WBS Excel文件为树形结构.

        Args:
            file_path: Excel文件路径

        Returns:
            Dict: WBS树形结构数据

        Raises:
            ValueError: 解析失败
        """
        import openpyxl

        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            if sheet is None:
                raise ValueError("Excel文件没有活动工作表")

            # 读取数据行
            rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] is None:  # 跳过空行
                    continue

                row_data = {
                    "row_idx": row_idx,
                    "level": int(row[0]) if row[0] else 1,
                    "name": str(row[1]) if row[1] else "",
                    "duration": int(row[2]) if row[2] else None,
                    "start_date": self._parse_excel_date(row[3]),
                    "end_date": self._parse_excel_date(row[4]),
                    "assignee_name": str(row[5]) if row[5] else None,
                    "deliverable": str(row[6]) if row[6] else None,
                    "predecessor_id": str(row[7]) if row[7] else None,
                }

                # 生成节点ID
                row_data["id"] = f"WBS-{row_idx}"

                rows.append(row_data)

            # 构建树形结构
            wbs_tree = self._build_wbs_tree(rows)

            return wbs_tree

        except Exception as e:
            logger.error(f"Failed to parse WBS Excel: {e}")
            raise ValueError(f"WBS Excel解析失败: {e}")

    def _parse_excel_date(
        self,
        value: Any,
    ) -> Optional[date]:
        """
        解析Excel日期值.

        Args:
            value: Excel单元格值

        Returns:
            Optional[date]: 解析后的日期或None
        """
        if value is None:
            return None

        if isinstance(value, date):
            return value

        if isinstance(value, str):
            try:
                return date.fromisoformat(value)
            except ValueError:
                # 尝试其他格式
                try:
                    return datetime.strptime(value, "%Y/%m/%d").date()
                except ValueError:
                    try:
                        return datetime.strptime(value, "%Y-%m-%d").date()
                    except ValueError:
                        return None

        return None

    def _build_wbs_tree(
        self,
        rows: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """
        从行数据构建WBS树形结构.

        Args:
            rows: Excel行数据列表

        Returns:
            Dict: WBS树形结构
        """
        if not rows:
            return {}

        # 创建根节点（项目级）
        root = WBSNode(
            id="WBS-ROOT",
            name="项目WBS",
            level=1,
        )

        # 构建节点映射和层级映射
        node_map: Dict[str, WBSNode] = {}
        level_stack: List[WBSNode] = [root]  # 当前层级栈

        for row in rows:
            node = WBSNode(
                id=row["id"],
                name=row["name"],
                level=row["level"],
                duration=row["duration"],
                start_date=row["start_date"],
                end_date=row["end_date"],
                assignee_name=row["assignee_name"],
                deliverable=row["deliverable"],
                dependencies=[row["predecessor_id"]] if row["predecessor_id"] else [],
            )

            node_map[node.id] = node

            # 找到正确的父节点
            while len(level_stack) > 1 and level_stack[-1].level >= node.level:
                level_stack.pop()

            parent = level_stack[-1]
            node.parent_id = parent.id
            parent.children.append(node)

            # 将当前节点加入栈
            level_stack.append(node)

        return root.to_dict()

    def _generate_wbs_excel(
        self,
        wbs_tree: Dict[str, Any],
        project_id: uuid.UUID,
    ) -> BytesIO:
        """
        生成WBS Excel文件.

        Args:
            wbs_tree: WBS树形结构数据
            project_id: 项目ID

        Returns:
            BytesIO: Excel文件流
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "WBS"

        # 设置列标题
        headers = [
            "层级", "任务名称", "工期(天)", "开始日期",
            "结束日期", "负责人", "交付物", "前置任务ID"
        ]

        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 扁平化树形结构为行数据
        rows_data = self._flatten_wbs_tree(wbs_tree)

        # 写入数据行
        for row_idx, row_data in enumerate(rows_data, start=2):
            sheet.cell(row=row_idx, column=1, value=row_data["level"])
            sheet.cell(row=row_idx, column=2, value=row_data["name"])
            sheet.cell(row=row_idx, column=3, value=row_data["duration"])
            
            if row_data["start_date"]:
                sheet.cell(row=row_idx, column=4, value=row_data["start_date"])
            if row_data["end_date"]:
                sheet.cell(row=row_idx, column=5, value=row_data["end_date"])
            
            sheet.cell(row=row_idx, column=6, value=row_data["assignee_name"])
            sheet.cell(row=row_idx, column=7, value=row_data["deliverable"])
            
            if row_data["dependencies"]:
                sheet.cell(row=row_idx, column=8, value=row_data["dependencies"][0])

        # 保存到内存缓冲区
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _flatten_wbs_tree(
        self,
        wbs_tree: Dict[str, Any],
        result: Optional[List[Dict[str, Any]]] = None,
    ) -> List[Dict[str, Any]]:
        """
        扁平化WBS树为行数据列表.

        Args:
            wbs_tree: WBS树形结构
            result: 结果列表（递归使用）

        Returns:
            List[Dict]: 扁平化的行数据列表
        """
        if result is None:
            result = []

        node = WBSNode.from_dict(wbs_tree)

        # 跳过根节点（项目级）
        if node.level > 1:
            result.append({
                "id": node.id,
                "level": node.level,
                "name": node.name,
                "duration": node.duration,
                "start_date": node.start_date.isoformat() if node.start_date else None,
                "end_date": node.end_date.isoformat() if node.end_date else None,
                "assignee_name": node.assignee_name,
                "assignee_id": node.assignee_id,
                "deliverable": node.deliverable,
                "dependencies": node.dependencies,
            })

        # 递归处理子节点
        for child_data in wbs_tree.get("children", []):
            self._flatten_wbs_tree(child_data, result)

        return result

    def _parse_wbs_sheet_data(
        self,
        sheet_data: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """
        解析飞书表格数据为WBS树形结构.

        Args:
            sheet_data: 飞书表格数据列表

        Returns:
            Dict: WBS树形结构
        """
        # 飞书表格数据格式与Excel类似
        rows = []
        for idx, row in enumerate(sheet_data, start=2):
            row_data = {
                "row_idx": idx,
                "level": int(row.get("层级", row.get("level", 1))),
                "name": row.get("任务名称", row.get("name", "")),
                "duration": int(row.get("工期(天)", row.get("duration", 0)) or 0),
                "start_date": self._parse_excel_date(
                    row.get("开始日期", row.get("start_date"))
                ),
                "end_date": self._parse_excel_date(
                    row.get("结束日期", row.get("end_date"))
                ),
                "assignee_name": row.get("负责人", row.get("assignee_name")),
                "deliverable": row.get("交付物", row.get("deliverable")),
                "predecessor_id": row.get("前置任务ID", row.get("predecessor_id")),
                "id": row.get("id", f"WBS-{idx}"),
            }
            rows.append(row_data)

        return self._build_wbs_tree(rows)

    def _count_nodes(
        self,
        wbs_tree: Dict[str, Any],
    ) -> int:
        """
        统计WBS树节点数量.

        Args:
            wbs_tree: WBS树形结构

        Returns:
            int: 节点总数
        """
        count = 1  # 当前节点
        for child in wbs_tree.get("children", []):
            count += self._count_nodes(child)
        return count