"""
PM Digital Employee - WBS API
项目经理数字员工系统 - WBS分解API端点

提供WBS导入导出、版本管理、任务同步、可视化数据接口。
"""

import os
import uuid
import tempfile
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, Query, Request, Body
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_db_session
from app.core.logging import get_logger
from app.domain.enums import DataSource
from app.services.wbs_service import WBSService

logger = get_logger(__name__)

# 创建WBS API路由
wbs_router = APIRouter(prefix="/api/v1/wbs", tags=["WBS Management"])


# ==================== Pydantic模型 ====================

class WBSNodeModel(BaseModel):
    """WBS节点请求模型."""

    id: Optional[str] = Field(None, description="节点ID")
    name: str = Field(..., description="节点名称")
    level: int = Field(..., ge=1, le=10, description="层级")
    parent_id: Optional[str] = Field(None, description="父节点ID")
    duration: Optional[int] = Field(None, ge=0, description="工期天数")
    start_date: Optional[str] = Field(None, description="开始日期（YYYY-MM-DD）")
    end_date: Optional[str] = Field(None, description="结束日期（YYYY-MM-DD）")
    assignee_id: Optional[str] = Field(None, description="负责人ID")
    assignee_name: Optional[str] = Field(None, description="负责人姓名")
    deliverable: Optional[str] = Field(None, description="交付物")
    dependencies: Optional[List[str]] = Field(default=[], description="前置依赖节点ID列表")
    description: Optional[str] = Field(None, description="节点描述")
    children: Optional[List["WBSNodeModel"]] = Field(default=[], description="子节点列表")


class CreateVersionRequest(BaseModel):
    """创建WBS版本请求."""

    project_id: str = Field(..., description="项目ID")
    wbs_data: WBSNodeModel = Field(..., description="WBS树形结构数据")
    version_name: Optional[str] = Field(None, description="版本名称")
    description: Optional[str] = Field(None, description="版本描述")
    data_source: Optional[str] = Field(default="lark_card", description="数据来源")


class PublishVersionRequest(BaseModel):
    """发布版本请求."""

    sync_to_tasks: bool = Field(default=True, description="是否同步到任务模块")


class RollbackVersionRequest(BaseModel):
    """回滚版本请求."""

    sync_to_tasks: bool = Field(default=True, description="是否同步到任务模块")


# ==================== WBS导入 ====================

@wbs_router.post("/import")
async def import_wbs(
    file: UploadFile = File(...),
    project_id: str = Query(..., description="项目ID"),
    version_name: Optional[str] = Query(None, description="版本名称"),
    description: Optional[str] = Query(None, description="版本描述"),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """
    导入WBS结构（Excel文件）.

    上传Excel文件，解析WBS树形结构，创建新版本。

    Args:
        file: 上传的Excel文件
        project_id: 项目ID
        version_name: 版本名称（可选）
        description: 版本描述（可选）
        session: 数据库会话

    Returns:
        Dict: 导入结果

    Raises:
        HTTPException: 参数错误或导入失败
    """
    # 项目ID解析
    try:
        project_uuid = uuid.UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="项目ID格式错误")

    # 文件类型校验
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="请上传Excel文件（.xlsx或.xls格式）",
        )

    try:
        # 保存上传文件到临时目录
        temp_dir = tempfile.mkdtemp()
        temp_file_path = os.path.join(temp_dir, file.filename)

        with open(temp_file_path, "wb") as f:
            content = await file.read()
            f.write(content)

        # 创建WBS服务
        service = WBSService(session)

        # 导入WBS
        version = await service.import_wbs_from_excel(
            file_path=temp_file_path,
            project_id=project_uuid,
            version_name=version_name,
            description=description,
        )

        # 清理临时文件
        os.remove(temp_file_path)
        os.rmdir(temp_dir)

        logger.info(
            f"WBS import completed: project_id={project_id}, "
            f"version_id={str(version.id)}, version_number={version.version_number}"
        )

        return {
            "code": 0,
            "msg": "WBS导入成功",
            "data": {
                "version_id": str(version.id),
                "version_number": version.version_number,
                "version_name": version.version_name,
                "status": version.status,
                "created_at": version.created_at.isoformat(),
            },
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Failed to import WBS: {e}")
        raise HTTPException(status_code=500, detail=f"WBS导入失败: {str(e)}")


# ==================== WBS导出 ====================

@wbs_router.get("/export/{project_id}")
async def export_wbs(
    project_id: str,
    version_id: Optional[str] = Query(None, description="版本ID（可选）"),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    """
    导出WBS为Excel.

    Args:
        project_id: 项目ID
        version_id: 版本ID（可选，默认使用当前版本）
        session: 数据库会话

    Returns:
        StreamingResponse: Excel文件流

    Raises:
        HTTPException: 参数错误或导出失败
    """
    # 项目ID解析
    try:
        project_uuid = uuid.UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="项目ID格式错误")

    # 版本ID解析
    version_uuid = None
    if version_id:
        try:
            version_uuid = uuid.UUID(version_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="版本ID格式错误")

    try:
        # 创建WBS服务
        service = WBSService(session)

        # 导出WBS
        buffer = await service.export_wbs_to_excel(
            project_id=project_uuid,
            version_id=version_uuid,
        )

        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"WBS_{project_id}_导出_{timestamp}.xlsx"

        logger.info(f"WBS export completed: project_id={project_id}, version_id={version_id}")

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
            },
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Failed to export WBS: {e}")
        raise HTTPException(status_code=500, detail=f"WBS导出失败: {str(e)}")


# ==================== 版本管理 ====================

@wbs_router.post("/version")
async def create_version(
    request: CreateVersionRequest,
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """
    创建新WBS版本.

    通过API直接提交WBS树形结构数据，创建新版本。

    Args:
        request: 创建版本请求
        session: 数据库会话

    Returns:
        Dict: 创建结果

    Raises:
        HTTPException: 参数错误或创建失败
    """
    # 项目ID解析
    try:
        project_uuid = uuid.UUID(request.project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="项目ID格式错误")

    # 数据来源解析
    try:
        data_source = DataSource(request.data_source)
    except ValueError:
        data_source = DataSource.LARK_CARD

    try:
        # 创建WBS服务
        service = WBSService(session)

        # 转换WBS数据为字典格式
        wbs_data_dict = request.wbs_data.dict()

        # 递归转换子节点
        if request.wbs_data.children:
            wbs_data_dict["children"] = [
                child.dict() for child in request.wbs_data.children
            ]

        # 创建版本
        version = await service.create_version(
            project_id=project_uuid,
            wbs_data=wbs_data_dict,
            version_name=request.version_name,
            description=request.description,
            data_source=data_source,
        )

        logger.info(
            f"WBS version created: project_id={request.project_id}, "
            f"version_id={str(version.id)}, version_number={version.version_number}"
        )

        return {
            "code": 0,
            "msg": "WBS版本创建成功",
            "data": {
                "version_id": str(version.id),
                "version_number": version.version_number,
                "version_name": version.version_name,
                "status": version.status,
                "is_published": version.is_published,
                "created_at": version.created_at.isoformat(),
            },
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Failed to create WBS version: {e}")
        raise HTTPException(status_code=500, detail=f"WBS版本创建失败: {str(e)}")


@wbs_router.get("/version/{project_id}")
async def get_version_list(
    project_id: str,
    include_draft: bool = Query(True, description="是否包含草稿版本"),
    limit: int = Query(50, ge=1, le=100, description="返回数量限制"),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """
    获取WBS版本列表.

    Args:
        project_id: 项目ID
        include_draft: 是否包含草稿版本
        limit: 返回数量限制
        session: 数据库会话

    Returns:
        Dict: 版本列表

    Raises:
        HTTPException: 参数错误或查询失败
    """
    # 项目ID解析
    try:
        project_uuid = uuid.UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="项目ID格式错误")

    try:
        # 创建WBS服务
        service = WBSService(session)

        # 获取版本历史
        versions = await service.get_version_history(
            project_id=project_uuid,
            include_draft=include_draft,
            limit=limit,
        )

        # 转换为响应格式
        version_list = []
        for version in versions:
            version_list.append({
                "version_id": str(version.id),
                "version_number": version.version_number,
                "version_name": version.version_name,
                "description": version.description,
                "status": version.status,
                "is_published": version.is_published,
                "is_current": version.is_current,
                "data_source": version.data_source,
                "created_by_name": version.created_by_name,
                "published_by_name": version.published_by_name,
                "published_at": version.published_at.isoformat() if version.published_at else None,
                "created_at": version.created_at.isoformat(),
            })

        return {
            "code": 0,
            "msg": "查询成功",
            "data": {
                "project_id": project_id,
                "total": len(version_list),
                "versions": version_list,
            },
        }

    except Exception as e:
        logger.error(f"Failed to get WBS versions: {e}")
        raise HTTPException(status_code=500, detail=f"查询失败: {str(e)}")


@wbs_router.post("/publish/{version_id}")
async def publish_version(
    version_id: str,
    project_id: str = Query(..., description="项目ID"),
    request: PublishVersionRequest = Body(default=PublishVersionRequest()),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """
    发布WBS版本.

    发布后该版本成为当前版本，可选自动同步到任务模块。

    Args:
        version_id: 版本ID
        project_id: 项目ID
        request: 发布请求参数
        session: 数据库会话

    Returns:
        Dict: 发布结果

    Raises:
        HTTPException: 参数错误或发布失败
    """
    # ID解析
    try:
        version_uuid = uuid.UUID(version_id)
        project_uuid = uuid.UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID格式错误")

    try:
        # 创建WBS服务
        service = WBSService(session)

        # 发布版本
        version = await service.publish_version(
            version_id=version_uuid,
            project_id=project_uuid,
            sync_to_tasks=request.sync_to_tasks,
        )

        logger.info(
            f"WBS version published: version_id={version_id}, "
            f"project_id={project_id}, sync_to_tasks={request.sync_to_tasks}"
        )

        return {
            "code": 0,
            "msg": "WBS版本发布成功",
            "data": {
                "version_id": str(version.id),
                "version_number": version.version_number,
                "version_name": version.version_name,
                "is_current": version.is_current,
                "published_at": version.published_at.isoformat(),
                "sync_to_tasks": request.sync_to_tasks,
            },
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Failed to publish WBS version: {e}")
        raise HTTPException(status_code=500, detail=f"WBS版本发布失败: {str(e)}")


@wbs_router.post("/rollback/{version_id}")
async def rollback_version(
    version_id: str,
    project_id: str = Query(..., description="项目ID"),
    request: RollbackVersionRequest = Body(default=RollbackVersionRequest()),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """
    回滚到指定WBS版本.

    Args:
        version_id: 要回滚到的版本ID
        project_id: 项目ID
        request: 回滚请求参数
        session: 数据库会话

    Returns:
        Dict: 回滚结果

    Raises:
        HTTPException: 参数错误或回滚失败
    """
    # ID解析
    try:
        version_uuid = uuid.UUID(version_id)
        project_uuid = uuid.UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID格式错误")

    try:
        # 创建WBS服务
        service = WBSService(session)

        # 回滚版本
        version = await service.rollback_to_version(
            version_id=version_uuid,
            project_id=project_uuid,
            sync_to_tasks=request.sync_to_tasks,
        )

        logger.info(
            f"WBS version rolled back: version_id={version_id}, "
            f"project_id={project_id}, sync_to_tasks={request.sync_to_tasks}"
        )

        return {
            "code": 0,
            "msg": "WBS版本回滚成功",
            "data": {
                "version_id": str(version.id),
                "version_number": version.version_number,
                "version_name": version.version_name,
                "is_current": version.is_current,
                "sync_to_tasks": request.sync_to_tasks,
            },
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Failed to rollback WBS version: {e}")
        raise HTTPException(status_code=500, detail=f"WBS版本回滚失败: {str(e)}")


# ==================== WBS可视化 ====================

@wbs_router.get("/tree/{project_id}")
async def get_wbs_tree(
    project_id: str,
    version_id: Optional[str] = Query(None, description="版本ID（可选）"),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """
    获取WBS树形结构.

    Args:
        project_id: 项目ID
        version_id: 版本ID（可选，默认使用当前版本）
        session: 数据库会话

    Returns:
        Dict: WBS树形结构数据

    Raises:
        HTTPException: 参数错误或查询失败
    """
    # ID解析
    try:
        project_uuid = uuid.UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="项目ID格式错误")

    version_uuid = None
    if version_id:
        try:
            version_uuid = uuid.UUID(version_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="版本ID格式错误")

    try:
        # 创建WBS服务
        service = WBSService(session)

        # 获取WBS树
        tree_data = await service.get_wbs_tree(
            project_id=project_uuid,
            version_id=version_uuid,
        )

        return {
            "code": 0,
            "msg": "查询成功",
            "data": tree_data,
        }

    except Exception as e:
        logger.error(f"Failed to get WBS tree: {e}")
        raise HTTPException(status_code=500, detail=f"查询失败: {str(e)}")


@wbs_router.get("/gantt/{project_id}")
async def get_gantt_data(
    project_id: str,
    version_id: Optional[str] = Query(None, description="版本ID（可选）"),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """
    获取甘特图数据.

    Args:
        project_id: 项目ID
        version_id: 版本ID（可选，默认使用当前版本）
        session: 数据库会话

    Returns:
        Dict: 甘特图数据

    Raises:
        HTTPException: 参数错误或查询失败
    """
    # ID解析
    try:
        project_uuid = uuid.UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="项目ID格式错误")

    version_uuid = None
    if version_id:
        try:
            version_uuid = uuid.UUID(version_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="版本ID格式错误")

    try:
        # 创建WBS服务
        service = WBSService(session)

        # 获取甘特图数据
        gantt_data = await service.get_gantt_data(
            project_id=project_uuid,
            version_id=version_uuid,
        )

        return {
            "code": 0,
            "msg": "查询成功",
            "data": gantt_data,
        }

    except Exception as e:
        logger.error(f"Failed to get Gantt data: {e}")
        raise HTTPException(status_code=500, detail=f"查询失败: {str(e)}")


# ==================== WBS同步到任务 ====================

@wbs_router.post("/sync/{version_id}")
async def sync_wbs_to_tasks(
    version_id: str,
    project_id: str = Query(..., description="项目ID"),
    clear_existing: bool = Query(False, description="是否清除现有任务"),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """
    手动同步WBS到任务模块.

    Args:
        version_id: WBS版本ID
        project_id: 项目ID
        clear_existing: 是否清除现有任务（用于回滚场景）
        session: 数据库会话

    Returns:
        Dict: 同步结果统计

    Raises:
        HTTPException: 参数错误或同步失败
    """
    # ID解析
    try:
        version_uuid = uuid.UUID(version_id)
        project_uuid = uuid.UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID格式错误")

    try:
        # 创建WBS服务
        service = WBSService(session)

        # 同步WBS到任务
        stats = await service.wbs_to_tasks(
            version_id=version_uuid,
            project_id=project_uuid,
            clear_existing=clear_existing,
        )

        logger.info(
            f"WBS synced to tasks: version_id={version_id}, "
            f"project_id={project_id}, tasks_created={stats['tasks_created']}"
        )

        return {
            "code": 0,
            "msg": "WBS同步成功",
            "data": {
                "version_id": version_id,
                "project_id": project_id,
                "clear_existing": clear_existing,
                "total_nodes": stats["total_nodes"],
                "tasks_created": stats["tasks_created"],
                "tasks_updated": stats["tasks_updated"],
                "errors_count": len(stats["errors"]),
                "errors": stats["errors"][:10] if stats["errors"] else [],
            },
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Failed to sync WBS to tasks: {e}")
        raise HTTPException(status_code=500, detail=f"WBS同步失败: {str(e)}")


# ==================== 当前版本查询 ====================

@wbs_router.get("/current/{project_id}")
async def get_current_version(
    project_id: str,
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """
    获取当前WBS版本.

    Args:
        project_id: 项目ID
        session: 数据库会话

    Returns:
        Dict: 当前版本信息

    Raises:
        HTTPException: 参数错误或查询失败
    """
    # 项目ID解析
    try:
        project_uuid = uuid.UUID(project_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="项目ID格式错误")

    try:
        # 创建WBS服务
        service = WBSService(session)

        # 获取当前版本
        version = await service.get_current_version(project_uuid)

        if not version:
            return {
                "code": 0,
                "msg": "项目暂无已发布的WBS版本",
                "data": {
                    "project_id": project_id,
                    "current_version": None,
                },
            }

        return {
            "code": 0,
            "msg": "查询成功",
            "data": {
                "project_id": project_id,
                "current_version": {
                    "version_id": str(version.id),
                    "version_number": version.version_number,
                    "version_name": version.version_name,
                    "description": version.description,
                    "status": version.status,
                    "is_published": version.is_published,
                    "is_current": version.is_current,
                    "data_source": version.data_source,
                    "published_by_name": version.published_by_name,
                    "published_at": version.published_at.isoformat() if version.published_at else None,
                    "created_at": version.created_at.isoformat(),
                },
            },
        }

    except Exception as e:
        logger.error(f"Failed to get current WBS version: {e}")
        raise HTTPException(status_code=500, detail=f"查询失败: {str(e)}")


# ==================== 模板下载 ====================

@wbs_router.get("/template")
async def download_wbs_template() -> StreamingResponse:
    """
    下载WBS导入模板.

    Returns:
        StreamingResponse: Excel模板文件流
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "WBS导入模板"

    # 设置列标题
    headers = [
        "层级", "任务名称", "工期(天)", "开始日期",
        "结束日期", "负责人", "交付物", "前置任务ID"
    ]

    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 添加示例数据
    sample_data = [
        [1, "项目名称", 30, "2024-01-01", "2024-01-31", "项目经理", "项目交付", None],
        [2, "需求分析", 5, "2024-01-01", "2024-01-05", "分析师A", "需求文档", None],
        [2, "设计阶段", 10, "2024-01-06", "2024-01-15", "设计师B", "设计文档", "WBS-2"],
        [3, "概要设计", 3, "2024-01-06", "2024-01-08", "设计师B", "概要设计文档", None],
        [3, "详细设计", 7, "2024-01-09", "2024-01-15", "设计师C", "详细设计文档", "WBS-4"],
        [2, "开发阶段", 15, "2024-01-16", "2024-01-30", "开发组", "系统代码", "WBS-3"],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # 添加说明
    sheet.cell(row=9, column=1, value="说明：")
    sheet.cell(row=10, column=1, value="1. 层级：1=项目级，2=任务级，3=子任务级...")
    sheet.cell(row=11, column=1, value="2. 前置任务ID：填写上一行的节点ID（如WBS-2），表示依赖关系")
    sheet.cell(row=12, column=1, value="3. 工期：以天为单位")
    sheet.cell(row=13, column=1, value="4. 日期格式：YYYY-MM-DD")

    # 保存到内存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"WBS导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
        },
    )