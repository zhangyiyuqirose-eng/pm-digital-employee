"""
PM Digital Employee - Excel Import Tasks
项目经理数字员工系统 - Excel异步导入Celery任务

提供异步Excel导入功能，避免大文件导入阻塞API请求。
"""

import os
import uuid
import tempfile
import json
from datetime import datetime
from typing import Any, Dict, List, Optional

from celery import shared_task
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.core.logging import get_logger
from app.domain.enums import ImportMode
from app.domain.models.excel_import_log import ExcelImportLog
from app.services.excel_service import ExcelService
from app.services.validation_service import ValidationService

logger = get_logger(__name__)

# 支持的模块
SUPPORTED_MODULES = ["project", "task", "milestone", "risk", "cost"]


def get_sync_session():
    """
    获取同步数据库会话（用于Celery任务）.

    Celery任务运行在独立进程，需要使用同步会话。

    Returns:
        Session: SQLAlchemy同步会话
    """
    # 创建同步引擎
    sync_url = settings.database_url.replace("+asyncpg", "")
    engine = create_engine(sync_url, echo=False)
    Session = sessionmaker(bind=engine)
    return Session()


@shared_task(
    name="excel.async_import",
    bind=True,
    max_retries=3,
    default_retry_delay=60,
)
def async_import_excel(
    self,
    file_path: str,
    module: str,
    import_mode: str,
    project_id: Optional[str] = None,
    user_id: Optional[str] = None,
) -> Dict[str, Any]:
    """
    异步导入Excel数据.

    Args:
        file_path: Excel文件临时路径
        module: 模块名称
        import_mode: 导入模式
        project_id: 项目ID（可选）
        user_id: 用户ID（可选）

    Returns:
        Dict: 导入结果
    """
    logger.info(
        f"Starting async Excel import: module={module}, mode={import_mode}, "
        f"file_path={file_path}, project_id={project_id}"
    )

    # 参数校验
    if module not in SUPPORTED_MODULES:
        return {
            "success": False,
            "error": f"不支持的模块 '{module}'",
        }

    try:
        # 创建同步数据库会话
        session = get_sync_session()

        try:
            # 创建Excel服务实例（使用同步适配）
            excel_service = SyncExcelServiceAdapter(session)

            # 解析Excel数据
            data_list, parse_errors = excel_service.parse_excel(file_path, module)

            if not data_list:
                return {
                    "success": True,
                    "message": "Excel文件中没有有效数据",
                    "rows_total": 0,
                    "rows_imported": 0,
                    "rows_failed": 0,
                }

            # 项目ID转换
            project_uuid = uuid.UUID(project_id) if project_id else None

            # 执行导入
            import_log = excel_service.import_data(
                data_list=data_list,
                module=module,
                import_mode=import_mode,
                project_id=project_uuid,
                user_id=user_id,
            )

            session.commit()

            result = {
                "success": True,
                "log_id": str(import_log.id),
                "rows_total": import_log.rows_total,
                "rows_imported": import_log.rows_imported,
                "rows_updated": import_log.rows_updated,
                "rows_failed": import_log.rows_failed,
                "validation_passed": import_log.validation_passed,
            }

            logger.info(f"Async Excel import completed: {result}")

            # 清理临时文件
            if os.path.exists(file_path):
                os.remove(file_path)
                # 尝试删除临时目录
                temp_dir = os.path.dirname(file_path)
                if temp_dir.startswith(tempfile.gettempdir()):
                    try:
                        os.rmdir(temp_dir)
                    except OSError:
                        pass  # 目录可能还有其他文件

            return result

        finally:
            session.close()

    except Exception as e:
        logger.error(f"Async Excel import failed: {e}")

        # 重试逻辑
        if self.request.retries < self.max_retries:
            logger.info(f"Retrying task, attempt {self.request.retries + 1}")
            raise self.retry(exc=e)

        return {
            "success": False,
            "error": str(e),
        }


@shared_task(
    name="excel.async_export",
    bind=True,
    max_retries=2,
    default_retry_delay=30,
)
def async_export_excel(
    self,
    module: str,
    project_id: Optional[str] = None,
    user_id: Optional[str] = None,
    output_dir: Optional[str] = None,
) -> Dict[str, Any]:
    """
    异步导出Excel数据.

    Args:
        module: 模块名称
        project_id: 项目ID（可选）
        user_id: 用户ID（可选）
        output_dir: 输出目录（可选）

    Returns:
        Dict: 导出结果，包含文件路径
    """
    logger.info(
        f"Starting async Excel export: module={module}, project_id={project_id}"
    )

    if module not in SUPPORTED_MODULES:
        return {
            "success": False,
            "error": f"不支持的模块 '{module}'",
        }

    try:
        session = get_sync_session()

        try:
            # 创建Excel服务实例
            excel_service = SyncExcelServiceAdapter(session)

            # 项目ID转换
            project_uuid = uuid.UUID(project_id) if project_id else None

            # 执行导出
            buffer = excel_service.export_data(module, project_uuid)

            # 确定输出目录
            if not output_dir:
                output_dir = tempfile.mkdtemp(prefix="excel_export_")

            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            if project_id:
                filename = f"{module}_{project_id}_导出_{timestamp}.xlsx"
            else:
                filename = f"{module}_全部导出_{timestamp}.xlsx"

            file_path = os.path.join(output_dir, filename)

            # 写入文件
            with open(file_path, "wb") as f:
                f.write(buffer.getvalue())

            logger.info(f"Async Excel export completed: file_path={file_path}")

            return {
                "success": True,
                "file_path": file_path,
                "filename": filename,
                "file_size": os.path.getsize(file_path),
            }

        finally:
            session.close()

    except Exception as e:
        logger.error(f"Async Excel export failed: {e}")

        if self.request.retries < self.max_retries:
            raise self.retry(exc=e)

        return {
            "success": False,
            "error": str(e),
        }


class SyncExcelServiceAdapter:
    """
    Excel服务同步适配器.

    将异步Excel服务适配为同步版本，用于Celery任务。
    """

    def __init__(self, session) -> None:
        """
        初始化适配器.

        Args:
            session: 同步数据库会话
        """
        self.session = session
        self.validation_service = ValidationService()

    def parse_excel(
        self,
        file_path: str,
        module: str,
    ) -> tuple:
        """
        解析Excel数据（同步版本）.

        Args:
            file_path: Excel文件路径
            module: 模块名称

        Returns:
            Tuple: (数据列表, 解析错误列表)
        """
        # 直接调用ExcelService的静态方法
        from app.core.validation_config import get_module_config
        from openpyxl import load_workbook

        config = get_module_config(module)
        if not config:
            raise ValueError(f"模块 '{module}' 不存在")

        # 加载Excel文件
        wb = load_workbook(file_path, data_only=True)
        ws = wb["数据"] if "数据" in wb.sheetnames else wb.active

        # 获取字段映射
        field_mapping = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=2, column=col).value
            if header:
                for field in config.fields:
                    if field.display_name == header:
                        field_mapping[col] = field.field_name
                        break

        # 解析数据行
        data_list: List[Dict[str, Any]] = []
        row_errors: List[Dict[str, Any]] = []

        for row_idx in range(4, ws.max_row + 1):
            row_data = {}
            is_row_valid = True

            for col in range(1, ws.max_column + 1):
                field_name = field_mapping.get(col)
                if not field_name:
                    continue

                value = ws.cell(row=row_idx, column=col).value

                if value is None and col == 1:
                    is_row_valid = False
                    break

                row_data[field_name] = value

            if is_row_valid and row_data:
                data_list.append(row_data)

        wb.close()
        return data_list, row_errors

    def import_data(
        self,
        data_list: List[Dict[str, Any]],
        module: str,
        import_mode: str,
        project_id: Optional[uuid.UUID] = None,
        user_id: Optional[str] = None,
    ) -> ExcelImportLog:
        """
        导入数据（同步版本）.

        Args:
            data_list: 数据列表
            module: 模块名称
            import_mode: 导入模式
            project_id: 项目ID
            user_id: 用户ID

        Returns:
            ExcelImportLog: 导入日志
        """
        from app.services.excel_service import TEMPLATE_VERSION

        # 创建导入日志
        import_log = ExcelImportLog(
            file_name=f"{module}_async_import_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
            import_mode=import_mode,
            template_version=TEMPLATE_VERSION,
            rows_total=len(data_list),
        )
        self.session.add(import_log)
        self.session.flush()

        # 执行校验
        validation_results = self.validation_service.validate_batch(data_list, module)

        valid_data = []
        row_errors = []

        for idx, result in enumerate(validation_results):
            if result.is_valid:
                valid_data.append(result.validated_data)
            else:
                row_errors.append({
                    "row_index": idx + 1,
                    "errors": result.errors,
                })

        import_log.validation_passed = len(row_errors) == 0
        import_log.validation_errors = json.dumps(row_errors[:100]) if row_errors else None
        import_log.rows_failed = len(row_errors)

        # 执行导入
        if valid_data:
            import_result = self._execute_import_sync(
                valid_data, module, import_mode, project_id
            )
            import_log.rows_imported = import_result.get("inserted", 0)
            import_log.rows_updated = import_result.get("updated", 0)
            import_log.rows_skipped = import_result.get("skipped", 0)

        import_log.row_errors = json.dumps(row_errors[:500]) if row_errors else None

        self.session.flush()
        self.session.refresh(import_log)

        return import_log

    def _execute_import_sync(
        self,
        data_list: List[Dict[str, Any]],
        module: str,
        import_mode: str,
        project_id: Optional[uuid.UUID] = None,
    ) -> Dict[str, int]:
        """
        执行同步导入操作.

        Args:
            data_list: 数据列表
            module: 模块名称
            import_mode: 导入模式
            project_id: 项目ID

        Returns:
            Dict: 导入统计
        """
        result = {"inserted": 0, "updated": 0, "skipped": 0}

        # 同步版本的导入逻辑
        if module == "project":
            from app.domain.models.project import Project
            for data in data_list:
                if import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                    existing_id = data.get("id")
                    if existing_id:
                        try:
                            existing = self.session.query(Project).filter(
                                Project.id == uuid.UUID(existing_id)
                            ).first()
                            if existing:
                                for key, value in data.items():
                                    setattr(existing, key, value)
                                result["updated"] += 1
                            else:
                                self.session.add(Project(**data))
                                result["inserted"] += 1
                        except Exception:
                            self.session.add(Project(**data))
                            result["inserted"] += 1
                    else:
                        self.session.add(Project(**data))
                        result["inserted"] += 1
                else:
                    self.session.add(Project(**data))
                    result["inserted"] += 1

        elif module == "task":
            from app.domain.models.task import Task
            target_project_id = project_id or data_list[0].get("project_id")
            for data in data_list:
                data["project_id"] = target_project_id
                if import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                    existing_id = data.get("id")
                    if existing_id:
                        try:
                            existing = self.session.query(Task).filter(
                                Task.id == uuid.UUID(existing_id),
                                Task.project_id == target_project_id
                            ).first()
                            if existing:
                                for key, value in data.items():
                                    if key != "project_id":
                                        setattr(existing, key, value)
                                result["updated"] += 1
                            else:
                                self.session.add(Task(**data))
                                result["inserted"] += 1
                        except Exception:
                            self.session.add(Task(**data))
                            result["inserted"] += 1
                    else:
                        self.session.add(Task(**data))
                        result["inserted"] += 1
                else:
                    self.session.add(Task(**data))
                    result["inserted"] += 1

        elif module == "milestone":
            from app.domain.models.milestone import Milestone
            target_project_id = project_id or data_list[0].get("project_id")
            for data in data_list:
                data["project_id"] = target_project_id
                if import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                    existing_id = data.get("id")
                    if existing_id:
                        try:
                            existing = self.session.query(Milestone).filter(
                                Milestone.id == uuid.UUID(existing_id),
                                Milestone.project_id == target_project_id
                            ).first()
                            if existing:
                                for key, value in data.items():
                                    if key != "project_id":
                                        setattr(existing, key, value)
                                result["updated"] += 1
                            else:
                                self.session.add(Milestone(**data))
                                result["inserted"] += 1
                        except Exception:
                            self.session.add(Milestone(**data))
                            result["inserted"] += 1
                    else:
                        self.session.add(Milestone(**data))
                        result["inserted"] += 1
                else:
                    self.session.add(Milestone(**data))
                    result["inserted"] += 1

        elif module == "risk":
            from app.domain.models.risk import ProjectRisk
            target_project_id = project_id or data_list[0].get("project_id")
            for data in data_list:
                data["project_id"] = target_project_id
                if import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                    existing_id = data.get("id")
                    if existing_id:
                        try:
                            existing = self.session.query(ProjectRisk).filter(
                                ProjectRisk.id == uuid.UUID(existing_id),
                                ProjectRisk.project_id == target_project_id
                            ).first()
                            if existing:
                                for key, value in data.items():
                                    if key != "project_id":
                                        setattr(existing, key, value)
                                result["updated"] += 1
                            else:
                                self.session.add(ProjectRisk(**data))
                                result["inserted"] += 1
                        except Exception:
                            self.session.add(ProjectRisk(**data))
                            result["inserted"] += 1
                    else:
                        self.session.add(ProjectRisk(**data))
                        result["inserted"] += 1
                else:
                    self.session.add(ProjectRisk(**data))
                    result["inserted"] += 1

        elif module == "cost":
            from app.domain.models.cost import ProjectCostBudget, ProjectCostActual
            target_project_id = project_id or data_list[0].get("project_id")
            cost_type = data_list[0].get("cost_type", "budget")
            for data in data_list:
                data["project_id"] = target_project_id
                if import_mode == ImportMode.INCREMENTAL_UPDATE.value:
                    existing_id = data.get("id")
                    if existing_id:
                        try:
                            model = ProjectCostBudget if cost_type == "budget" else ProjectCostActual
                            existing = self.session.query(model).filter(
                                model.id == uuid.UUID(existing_id),
                                model.project_id == target_project_id
                            ).first()
                            if existing:
                                for key, value in data.items():
                                    if key != "project_id":
                                        setattr(existing, key, value)
                                result["updated"] += 1
                            else:
                                self.session.add(model(**data))
                                result["inserted"] += 1
                        except Exception:
                            model = ProjectCostBudget if cost_type == "budget" else ProjectCostActual
                            self.session.add(model(**data))
                            result["inserted"] += 1
                    else:
                        model = ProjectCostBudget if cost_type == "budget" else ProjectCostActual
                        self.session.add(model(**data))
                        result["inserted"] += 1
                else:
                    model = ProjectCostBudget if cost_type == "budget" else ProjectCostActual
                    self.session.add(model(**data))
                    result["inserted"] += 1

        self.session.flush()
        return result

    def export_data(
        self,
        module: str,
        project_id: Optional[uuid.UUID] = None,
    ):
        """
        导出数据（同步版本）.

        Args:
            module: 模块名称
            project_id: 项目ID

        Returns:
            BytesIO: Excel文件流
        """
        # 使用异步服务的模板生成功能（不需要数据库）
        from io import BytesIO
        from openpyxl import Workbook
        from app.services.excel_service import ExcelService

        # 创建临时服务实例用于模板生成
        temp_service = ExcelService(None)
        return temp_service.export_data.__wrapped__(self, module, project_id)  # 调用原始同步方法