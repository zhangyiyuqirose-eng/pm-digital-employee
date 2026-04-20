"""
PM Digital Employee - Excel Service Test Script
项目经理数字员工系统 - Excel服务测试脚本

用于验证Excel导入导出功能的基本工作。
"""

import os
import sys
import tempfile
from io import BytesIO

# 确保可以导入应用模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.services.excel_service import ExcelService


def test_template_generation():
    """测试模板生成功能."""
    print("=" * 60)
    print("测试1: Excel模板生成")
    print("=" * 60)

    service = ExcelService(None)

    modules = ["project", "task", "milestone", "risk", "cost"]

    for module in modules:
        try:
            buffer = service.generate_template(module)
            print(f"  ✅ {module} 模板生成成功 (大小: {buffer.getbuffer().nbytes} bytes)")

            # 保存模板到临时文件
            temp_dir = tempfile.mkdtemp()
            file_path = os.path.join(temp_dir, f"{module}_template.xlsx")
            with open(file_path, "wb") as f:
                f.write(buffer.getvalue())

            # 验证可以读取
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            sheets = wb.sheetnames
            print(f"    工作表: {sheets}")
            wb.close()

            # 清理
            os.remove(file_path)
            os.rmdir(temp_dir)

        except Exception as e:
            print(f"  ❌ {module} 模板生成失败: {e}")

    print()


def test_parse_excel():
    """测试Excel解析功能."""
    print("=" * 60)
    print("测试2: Excel数据解析")
    print("=" * 60)

    service = ExcelService(None)

    # 生成模板并添加测试数据
    buffer = service.generate_template("project")
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, "project_test.xlsx")

    with open(file_path, "wb") as f:
        f.write(buffer.getvalue())

    # 使用openpyxl添加测试数据
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb["数据"]

    # 添加测试数据行（从第4行开始）
    ws.cell(row=4, column=1, value="测试项目1")
    ws.cell(row=4, column=2, value="项目名称")
    ws.cell(row=4, column=3, value="研发项目")
    ws.cell(row=4, column=4, value="2025-01-01")
    ws.cell(row=4, column=5, value="2025-12-31")
    ws.cell(row=4, column=6, value="100000")

    ws.cell(row=5, column=1, value="测试项目2")
    ws.cell(row=5, column=2, value="项目名称2")
    ws.cell(row=5, column=3, value="运维项目")
    ws.cell(row=5, column=4, value="2025-02-01")
    ws.cell(row=5, column=5, value="2025-06-30")
    ws.cell(row=5, column=6, value="50000")

    wb.save(file_path)
    wb.close()

    # 解析Excel
    try:
        data_list, errors = service.parse_excel(file_path, "project")
        print(f"  ✅ 解析成功，数据行数: {len(data_list)}")
        if data_list:
            print(f"    第一行数据: {data_list[0]}")
        if errors:
            print(f"    解析错误: {errors}")
    except Exception as e:
        print(f"  ❌ 解析失败: {e}")

    # 清理
    os.remove(file_path)
    os.rmdir(temp_dir)

    print()


def test_validation_service():
    """测试校验服务."""
    print("=" * 60)
    print("测试3: 数据校验")
    print("=" * 60)

    from app.services.validation_service import ValidationService

    validation_service = ValidationService()

    # 测试数据
    test_data = {
        "name": "测试项目",
        "project_type": "研发项目",
        "start_date": "2025-01-01",
        "end_date": "2025-12-31",
        "total_budget": "100000",
    }

    try:
        result = validation_service.validate_all(test_data, "project")
        print(f"  ✅ 校验结果: is_valid={result.is_valid}")
        print(f"    错误数: {len(result.errors)}")
        if result.errors:
            print(f"    错误详情: {result.errors}")
        if result.validated_data:
            print(f"    校验后数据: {result.validated_data}")
    except Exception as e:
        print(f"  ❌ 校验失败: {e}")

    # 测试无效数据
    invalid_data = {
        "project_type": "无效类型",  # 枚举值不在允许列表中
        "total_budget": "-100",     # 金额小于0
    }

    try:
        result = validation_service.validate_all(invalid_data, "project")
        print(f"  ✅ 无效数据校验: is_valid={result.is_valid}")
        print(f"    错误数: {len(result.errors)}")
        if result.errors:
            for error in result.errors[:3]:
                print(f"    - {error.get('message', error)}")
    except Exception as e:
        print(f"  ❌ 校验失败: {e}")

    print()


def main():
    """运行所有测试."""
    print("\n" + "=" * 60)
    print("PM Digital Employee - Excel服务测试")
    print("=" * 60 + "\n")

    test_template_generation()
    test_parse_excel()
    test_validation_service()

    print("=" * 60)
    print("测试完成!")
    print("=" * 60)


if __name__ == "__main__":
    main()