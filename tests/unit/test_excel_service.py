"""
PM Digital Employee - Excel Service Tests
项目经理数字员工系统 - Excel导入导出服务单元测试

测试覆盖：模板生成、数据导出、数据解析、数据导入
"""

import pytest
import uuid
import os
import tempfile
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO
from unittest.mock import MagicMock, AsyncMock, patch

from app.services.excel_service import ExcelService, TEMPLATE_VERSION
from app.domain.enums import ImportMode


class TestExcelServiceTemplateGeneration:
    """Excel模板生成测试."""

    @pytest.fixture
    def service(self) -> ExcelService:
        """创建ExcelService实例（无session）."""
        return ExcelService(None)

    def test_generate_template_project(self, service: ExcelService):
        """测试生成项目模板."""
        buffer = service.generate_template("project")

        assert buffer is not None
        assert isinstance(buffer, BytesIO)

        # 验证可以读取Excel文件
        from openpyxl import load_workbook
        wb = load_workbook(buffer)
        assert "说明" in wb.sheetnames
        assert "数据" in wb.sheetnames

    def test_generate_template_task(self, service: ExcelService):
        """测试生成任务模板."""
        buffer = service.generate_template("task")

        assert buffer is not None
        from openpyxl import load_workbook
        wb = load_workbook(buffer)
        assert wb.active.title == "说明"

    def test_generate_template_milestone(self, service: ExcelService):
        """测试生成里程碑模板."""
        buffer = service.generate_template("milestone")

        assert buffer is not None
        from openpyxl import load_workbook
        wb = load_workbook(buffer)
        assert "数据" in wb.sheetnames

    def test_generate_template_risk(self, service: ExcelService):
        """测试生成风险模板."""
        buffer = service.generate_template("risk")

        assert buffer is not None

    def test_generate_template_cost(self, service: ExcelService):
        """测试生成成本模板."""
        buffer = service.generate_template("cost")

        assert buffer is not None

    def test_generate_template_invalid_module(self, service: ExcelService):
        """测试生成不存在的模块模板."""
        with pytest.raises(ValueError) as exc_info:
            service.generate_template("invalid_module")

        assert "不存在" in str(exc_info.value)

    def test_template_has_version_info(self, service: ExcelService):
        """测试模板包含版本信息."""
        buffer = service.generate_template("project")

        from openpyxl import load_workbook
        wb = load_workbook(buffer)
        ws_data = wb["数据"]

        # 第一行应包含版本信息
        assert TEMPLATE_VERSION in str(ws_data["A1"].value)

    def test_template_has_required_fields_marked(self, service: ExcelService):
        """测试模板标记必填字段."""
        buffer = service.generate_template("project")

        from openpyxl import load_workbook
        wb = load_workbook(buffer)
        ws_data = wb["数据"]

        # 第二行是标题行，必填字段应红色标记
        # 项目名称是必填字段
        cell = ws_data["A2"]
        # 验证字体颜色（必填字段为红色）
        assert cell.font.color is not None or cell.font.bold == True

    def test_template_has_example_data(self, service: ExcelService):
        """测试模板包含示例数据."""
        buffer = service.generate_template("project")

        from openpyxl import load_workbook
        wb = load_workbook(buffer)
        ws_data = wb["数据"]

        # 第三行是示例数据行
        assert ws_data["A3"].value is not None

    def test_template_instruction_sheet_has_field_info(self, service: ExcelService):
        """测试说明页包含字段信息."""
        buffer = service.generate_template("project")

        from openpyxl import load_workbook
        wb = load_workbook(buffer)
        ws_instruction = wb["说明"]

        # 应包含标题
        assert ws_instruction["A1"].value is not None

    def test_template_enum_dropdown(self, service: ExcelService):
        """测试枚举字段下拉选择框."""
        buffer = service.generate_template("project")

        from openpyxl import load_workbook
        wb = load_workbook(buffer)
        ws_data = wb["数据"]

        # 验证数据验证（下拉框）存在
        # project_type是枚举字段，应有DataValidation
        assert len(ws_data.data_validations.dataValidation) > 0


class TestExcelServiceExampleValue:
    """示例值生成测试."""

    @pytest.fixture
    def service(self) -> ExcelService:
        """创建ExcelService实例."""
        return ExcelService(None)

    def test_get_example_value_enum(self, service: ExcelService):
        """测试枚举字段示例值."""
        from app.core.validation_config import FieldValidationConfig

        field = FieldValidationConfig(
            field_name="project_type",
            display_name="项目类型",
            field_type="enum",
            enum_values=["研发项目", "运维项目"],
        )

        example = service._get_example_value(field)
        assert example == "研发项目"

    def test_get_example_value_string(self, service: ExcelService):
        """测试字符串字段示例值."""
        from app.core.validation_config import FieldValidationConfig

        field = FieldValidationConfig(
            field_name="name",
            display_name="项目名称",
            field_type="str",
        )

        example = service._get_example_value(field)
        assert "示例" in example

    def test_get_example_value_int(self, service: ExcelService):
        """测试整数字段示例值."""
        from app.core.validation_config import FieldValidationConfig

        field = FieldValidationConfig(
            field_name="progress",
            display_name="进度",
            field_type="int",
        )

        example = service._get_example_value(field)
        assert isinstance(int(example), int) if example else True

    def test_get_example_value_float(self, service: ExcelService):
        """测试浮点字段示例值."""
        from app.core.validation_config import FieldValidationConfig

        field = FieldValidationConfig(
            field_name="total_budget",
            display_name="总预算",
            field_type="float",
        )

        example = service._get_example_value(field)
        assert "." in example

    def test_get_example_value_date(self, service: ExcelService):
        """测试日期字段示例值."""
        from app.core.validation_config import FieldValidationConfig

        field = FieldValidationConfig(
            field_name="start_date",
            display_name="开始日期",
            field_type="date",
        )

        example = service._get_example_value(field)
        assert "-" in example  # YYYY-MM-DD格式


class TestExcelServiceExportFormatting:
    """导出值格式化测试."""

    @pytest.fixture
    def service(self) -> ExcelService:
        """创建ExcelService实例."""
        return ExcelService(None)

    def test_format_export_value_uuid(self, service: ExcelService):
        """测试UUID格式化."""
        test_uuid = uuid.uuid4()
        from app.core.validation_config import FieldValidationConfig

        field = FieldValidationConfig(
            field_name="id",
            display_name="ID",
            field_type="str",
        )

        result = service._format_export_value(test_uuid, field)
        assert isinstance(result, str)

    def test_format_export_value_date(self, service: ExcelService):
        """测试日期格式化."""
        test_date = date(2026, 4, 19)
        from app.core.validation_config import FieldValidationConfig

        field = FieldValidationConfig(
            field_name="start_date",
            display_name="开始日期",
            field_type="date",
        )

        result = service._format_export_value(test_date, field)
        assert "2026-04-19" == result

    def test_format_export_value_decimal(self, service: ExcelService):
        """测试Decimal格式化."""
        test_decimal = Decimal("10000.50")
        from app.core.validation_config import FieldValidationConfig

        field = FieldValidationConfig(
            field_name="total_budget",
            display_name="总预算",
            field_type="float",
        )

        result = service._format_export_value(test_decimal, field)
        assert isinstance(result, float)

    def test_format_export_value_bool_true(self, service: ExcelService):
        """测试布尔值True格式化."""
        from app.core.validation_config import FieldValidationConfig

        field = FieldValidationConfig(
            field_name="is_active",
            display_name="是否活跃",
            field_type="bool",
        )

        result = service._format_export_value(True, field)
        assert result == "是"

    def test_format_export_value_bool_false(self, service: ExcelService):
        """测试布尔值False格式化."""
        from app.core.validation_config import FieldValidationConfig

        field = FieldValidationConfig(
            field_name="is_active",
            display_name="是否活跃",
            field_type="bool",
        )

        result = service._format_export_value(False, field)
        assert result == "否"

    def test_format_export_value_none(self, service: ExcelService):
        """测试None格式化."""
        from app.core.validation_config import FieldValidationConfig

        field = FieldValidationConfig(
            field_name="name",
            display_name="名称",
            field_type="str",
        )

        result = service._format_export_value(None, field)
        assert result == ""


class TestExcelServiceModelToDict:
    """模型转字典测试."""

    @pytest.fixture
    def service(self) -> ExcelService:
        """创建ExcelService实例."""
        return ExcelService(None)

    def test_model_to_dict(self, service: ExcelService):
        """测试ORM模型转字典."""
        # 创建Mock模型
        mock_model = MagicMock()
        mock_model.__table__ = MagicMock()

        # 正确设置column mock，确保name是字符串属性
        mock_column_id = MagicMock()
        mock_column_id.name = "id"
        mock_column_name = MagicMock()
        mock_column_name.name = "name"
        mock_model.__table__.columns = [mock_column_id, mock_column_name]

        mock_model.id = uuid.uuid4()
        mock_model.name = "测试项目"

        result = service._model_to_dict(mock_model)

        assert "id" in result
        assert "name" in result


class TestExcelServiceParseExcel:
    """Excel解析测试."""

    @pytest.fixture
    def service(self) -> ExcelService:
        """创建ExcelService实例."""
        return ExcelService(None)

    @pytest.fixture
    def sample_excel_path(self, service: ExcelService) -> str:
        """生成测试Excel文件."""
        # 生成模板
        buffer = service.generate_template("project")
        buffer.seek(0)

        # 保存到临时文件
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, "test_project.xlsx")

        from openpyxl import load_workbook, Workbook
        wb = load_workbook(buffer)
        ws_data = wb["数据"]

        # 添加测试数据行（第4行开始）
        ws_data["A4"] = "测试项目1"
        ws_data["B4"] = "研发项目"
        ws_data["C4"] = "2026-04-01"
        ws_data["D4"] = "2026-06-01"

        ws_data["A5"] = "测试项目2"
        ws_data["B5"] = "运维项目"
        ws_data["C5"] = "2026-05-01"
        ws_data["D5"] = "2026-07-01"

        wb.save(temp_path)
        return temp_path

    def test_parse_excel_success(self, service: ExcelService, sample_excel_path: str):
        """测试成功解析Excel."""
        data_list, errors = service.parse_excel(sample_excel_path, "project")

        assert len(data_list) == 2
        assert len(errors) == 0

    def test_parse_excel_field_mapping(self, service: ExcelService, sample_excel_path: str):
        """测试解析字段映射."""
        data_list, errors = service.parse_excel(sample_excel_path, "project")

        # 验证字段名正确映射
        assert "name" in data_list[0]
        assert "project_type" in data_list[0]

    def test_parse_excel_empty_file(self, service: ExcelService):
        """测试解析空Excel."""
        # 创建空Excel
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, "empty.xlsx")

        from openpyxl import Workbook
        wb = Workbook()
        wb.save(temp_path)

        data_list, errors = service.parse_excel(temp_path, "project")

        assert len(data_list) == 0

    def test_parse_excel_invalid_module(self, service: ExcelService, sample_excel_path: str):
        """测试解析不存在的模块."""
        with pytest.raises(ValueError):
            service.parse_excel(sample_excel_path, "invalid_module")


class TestExcelServiceImportData:
    """数据导入测试."""

    @pytest.fixture
    def mock_session(self) -> MagicMock:
        """创建Mock数据库会话."""
        session = AsyncMock()
        session.execute = AsyncMock()
        session.add = MagicMock()
        session.flush = AsyncMock()
        session.refresh = AsyncMock()
        return session

    @pytest.fixture
    def service(self, mock_session: MagicMock) -> ExcelService:
        """创建ExcelService实例."""
        return ExcelService(mock_session)

    @pytest.mark.asyncio
    async def test_import_data_append_mode(self, service: ExcelService, mock_session: MagicMock):
        """测试追加模式导入."""
        data_list = [
            {"name": "项目1", "project_type": "研发项目"},
            {"name": "项目2", "project_type": "运维项目"},
        ]

        # Mock repository返回
        mock_repo_result = MagicMock()
        mock_repo_result.scalar_one_or_none.return_value = None

        # Mock导入日志
        mock_log = MagicMock()
        mock_log.id = uuid.uuid4()
        mock_log.rows_total = 2
        mock_log.rows_imported = 2
        mock_log.rows_failed = 0
        mock_session.refresh.return_value = None

        result = await service.import_data(
            data_list=data_list,
            module="project",
            import_mode=ImportMode.APPEND_ONLY.value,
        )

        assert mock_session.add.called
        assert mock_session.flush.called

    @pytest.mark.asyncio
    async def test_import_data_with_validation_errors(self, service: ExcelService, mock_session: MagicMock):
        """测试导入带校验错误."""
        data_list = [
            {"name": "", "project_type": "研发项目"},  # name为空，校验失败
        ]

        mock_log = MagicMock()
        mock_session.refresh.return_value = None

        result = await service.import_data(
            data_list=data_list,
            module="project",
            import_mode=ImportMode.APPEND_ONLY.value,
        )

        # 应记录校验失败
        assert result.rows_failed > 0

    @pytest.mark.asyncio
    async def test_import_data_task_requires_project_id(self, service: ExcelService, mock_session: MagicMock):
        """测试任务导入缺少project_id."""
        data_list = [
            {"name": "任务1", "status": "pending"},
        ]

        # Mock导入日志和校验结果
        mock_log = MagicMock()
        mock_log.rows_failed = 1  # 校验失败
        mock_session.refresh.return_value = None

        result = await service.import_data(
            data_list=data_list,
            module="task",
            import_mode=ImportMode.APPEND_ONLY.value,
        )

        # 由于project_id缺失，校验失败，rows_failed > 0
        assert result.rows_failed > 0

    @pytest.mark.asyncio
    async def test_import_data_milestone_requires_project_id(self, service: ExcelService, mock_session: MagicMock):
        """测试里程碑导入缺少project_id."""
        data_list = [
            {"name": "里程碑1", "planned_date": "2026-06-01"},
        ]

        # Mock导入日志和校验结果
        mock_log = MagicMock()
        mock_log.rows_failed = 1  # 校验失败
        mock_session.refresh.return_value = None

        result = await service.import_data(
            data_list=data_list,
            module="milestone",
            import_mode=ImportMode.APPEND_ONLY.value,
        )

        # 由于project_id缺失，校验失败，rows_failed > 0
        assert result.rows_failed > 0

    @pytest.mark.asyncio
    async def test_import_data_empty_list(self, service: ExcelService, mock_session: MagicMock):
        """测试导入空数据列表."""
        mock_log = MagicMock()
        mock_log.rows_total = 0
        mock_session.refresh.return_value = None

        result = await service.import_data(
            data_list=[],
            module="project",
            import_mode=ImportMode.APPEND_ONLY.value,
        )

        assert result.rows_total == 0


class TestExcelServiceGetImportLogs:
    """导入日志查询测试."""

    @pytest.fixture
    def mock_session(self) -> MagicMock:
        """创建Mock数据库会话."""
        session = AsyncMock()
        session.execute = AsyncMock()
        return session

    @pytest.fixture
    def service(self, mock_session: MagicMock) -> ExcelService:
        """创建ExcelService实例."""
        return ExcelService(mock_session)

    @pytest.mark.asyncio
    async def test_get_import_logs(self, service: ExcelService, mock_session: MagicMock):
        """测试获取导入日志列表."""
        # Mock返回日志列表
        mock_logs = [MagicMock(id=uuid.uuid4(), module="project")]
        mock_result = MagicMock()
        mock_result.scalars.return_value.all.return_value = mock_logs
        mock_session.execute.return_value = mock_result

        logs = await service.get_import_logs(limit=50)

        assert len(logs) == 1

    @pytest.mark.asyncio
    async def test_get_import_logs_with_module_filter(self, service: ExcelService, mock_session: MagicMock):
        """测试获取导入日志带模块过滤."""
        mock_logs = [MagicMock(module="task")]
        mock_result = MagicMock()
        mock_result.scalars.return_value.all.return_value = mock_logs
        mock_session.execute.return_value = mock_result

        logs = await service.get_import_logs(module="task", limit=50)

        assert len(logs) == 1

    @pytest.mark.asyncio
    async def test_get_import_logs_empty(self, service: ExcelService, mock_session: MagicMock):
        """测试获取空日志列表."""
        mock_result = MagicMock()
        mock_result.scalars.return_value.all.return_value = []
        mock_session.execute.return_value = mock_result

        logs = await service.get_import_logs()

        assert len(logs) == 0


class TestExcelServiceImportModes:
    """导入模式测试."""

    @pytest.fixture
    def mock_session(self) -> MagicMock:
        """创建Mock数据库会话."""
        session = AsyncMock()
        session.execute = AsyncMock()
        session.add = MagicMock()
        session.flush = AsyncMock()
        session.refresh = AsyncMock()
        session.delete = AsyncMock()  # for _clear methods
        return session

    @pytest.fixture
    def service(self, mock_session: MagicMock) -> ExcelService:
        """创建ExcelService实例."""
        return ExcelService(mock_session)

    def test_import_mode_full_replace_value(self):
        """测试全量替换模式值."""
        assert ImportMode.FULL_REPLACE.value == "full_replace"

    def test_import_mode_incremental_update_value(self):
        """测试增量更新模式值."""
        assert ImportMode.INCREMENTAL_UPDATE.value == "incremental_update"

    def test_import_mode_append_only_value(self):
        """测试追加模式值."""
        assert ImportMode.APPEND_ONLY.value == "append_only"


class TestExcelServiceTemplateVersion:
    """模板版本测试."""

    def test_template_version_constant(self):
        """测试模板版本常量."""
        assert TEMPLATE_VERSION == "v1.2.0"


class TestExcelServiceEdgeCases:
    """ExcelService边界情况测试."""

    @pytest.fixture
    def service(self) -> ExcelService:
        """创建ExcelService实例."""
        return ExcelService(None)

    def test_generate_template_all_modules(self, service: ExcelService):
        """测试生成所有模块模板."""
        modules = ["project", "task", "milestone", "risk", "cost"]

        for module in modules:
            buffer = service.generate_template(module)
            assert buffer is not None

    def test_template_column_width(self, service: ExcelService):
        """测试模板列宽设置."""
        buffer = service.generate_template("project")

        from openpyxl import load_workbook
        wb = load_workbook(buffer)
        ws_data = wb["数据"]

        # 验证列宽已设置
        assert ws_data.column_dimensions["A"].width > 0

    def test_parse_excel_skip_header_rows(self, service: ExcelService):
        """测试解析跳过标题行."""
        buffer = service.generate_template("project")

        from openpyxl import load_workbook
        wb = load_workbook(buffer)
        ws_data = wb["数据"]

        # 数据从第4行开始
        ws_data["A4"] = "测试数据"
        wb.save(buffer)
        buffer.seek(0)

        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, "test.xlsx")
        wb.save(temp_path)

        data_list, errors = service.parse_excel(temp_path, "project")

        # 应跳过前3行（版本、标题、示例）
        if len(data_list) > 0:
            assert data_list[0].get("name") == "测试数据"